"""
Utility functions for registry app
"""
from datetime import datetime, date


def thai_month_name(month, short=False):
    """
    แปลงเลขเดือนเป็นชื่อเดือนภาษาไทย

    Args:
        month (int): เลขเดือน (1-12)
        short (bool): ใช้ชื่อย่อหรือไม่ (ธ.ค. vs ธันวาคม)

    Returns:
        str: ชื่อเดือนภาษาไทย
    """
    months_full = {
        1: 'มกราคม',
        2: 'กุมภาพันธ์',
        3: 'มีนาคม',
        4: 'เมษายน',
        5: 'พฤษภาคม',
        6: 'มิถุนายน',
        7: 'กรกฎาคม',
        8: 'สิงหาคม',
        9: 'กันยายน',
        10: 'ตุลาคม',
        11: 'พฤศจิกายน',
        12: 'ธันวาคม'
    }

    months_short = {
        1: 'ม.ค.',
        2: 'ก.พ.',
        3: 'มี.ค.',
        4: 'เม.ย.',
        5: 'พ.ค.',
        6: 'มิ.ย.',
        7: 'ก.ค.',
        8: 'ส.ค.',
        9: 'ก.ย.',
        10: 'ต.ค.',
        11: 'พ.ย.',
        12: 'ธ.ค.'
    }

    if short:
        return months_short.get(month, '')
    return months_full.get(month, '')


def arabic_to_thai_numerals(number):
    """
    แปลงเลขอารบิกเป็นเลขไทย

    Args:
        number: เลขอารบิก (int หรือ str)

    Returns:
        str: เลขไทย
    """
    thai_digits = ['๐', '๑', '๒', '๓', '๔', '๕', '๖', '๗', '๘', '๙']
    number_str = str(number)
    return ''.join(thai_digits[int(digit)] for digit in number_str)


def format_thai_date(date_value, short=False, use_thai_numerals=False):
    """
    แปลงวันที่เป็นรูปแบบไทย

    Args:
        date_value: วันที่ (date object, datetime object, หรือ string 'YYYY-MM-DD')
        short (bool): ใช้รูปแบบสั้นหรือไม่
        use_thai_numerals (bool): ใช้เลขไทยหรือไม่

    Returns:
        str: วันที่ภาษาไทย
        - แบบยาว: "24 ธันวาคม 2568" หรือ "๒๔ ธันวาคม ๒๕๖๘"
        - แบบสั้น: "24 ธ.ค. 2568" หรือ "๒๔ ธ.ค. ๒๕๖๘"

    Examples:
        >>> format_thai_date('2025-12-24')
        '24 ธันวาคม 2568'
        >>> format_thai_date('2025-12-24', short=True, use_thai_numerals=True)
        '๒๔ ธ.ค. ๒๕๖๘'
    """
    if not date_value:
        return ''

    # Convert string to date object if needed
    if isinstance(date_value, str):
        try:
            date_value = datetime.strptime(date_value, '%Y-%m-%d').date()
        except ValueError:
            return ''

    # Convert datetime to date if needed
    if isinstance(date_value, datetime):
        date_value = date_value.date()

    if not isinstance(date_value, date):
        return ''

    # แปลงเป็น พ.ศ. (เพิ่ม 543 ปี)
    day = date_value.day
    month = thai_month_name(date_value.month, short=short)
    year = date_value.year + 543

    # แปลงเป็นเลขไทยถ้าต้องการ
    if use_thai_numerals:
        day = arabic_to_thai_numerals(day)
        year = arabic_to_thai_numerals(year)

    return f"{day} {month} {year}"


def format_thai_date_range(start_date, end_date, short=False):
    """
    แปลงช่วงวันที่เป็นรูปแบบไทย

    Args:
        start_date: วันที่เริ่มต้น
        end_date: วันที่สิ้นสุด
        short (bool): ใช้รูปแบบสั้นหรือไม่

    Returns:
        str: ช่วงวันที่ภาษาไทย

    Examples:
        >>> format_thai_date_range('2025-12-01', '2025-12-15')
        '1 ธันวาคม 2568 - 15 ธันวาคม 2568'
        >>> format_thai_date_range('2025-12-01', '2025-12-15', short=True)
        '1 ธ.ค. 2568 - 15 ธ.ค. 2568'
    """
    start_str = format_thai_date(start_date, short=short)
    end_str = format_thai_date(end_date, short=short)

    if start_str and end_str:
        return f"{start_str} - {end_str}"
    elif start_str:
        return start_str
    elif end_str:
        return end_str
    return ''


def detect_excel_format(ws):
    """
    ตรวจสอบว่า Excel file เป็นฟอร์มเก่าหรือฟอร์มใหม่

    ฟอร์มใหม่: มี 3 คอลัมน์แยก (ยศ | ชื่อ | นามสกุล)
    ฟอร์มเก่า: รวม 1 คอลัมน์ (ยศ - ชื่อ สกุล)

    Args:
        ws: openpyxl worksheet object

    Returns:
        str: 'NEW' หรือ 'OLD'
    """
    try:
        # อ่าน header row (row 3)
        header_row = list(ws.iter_rows(min_row=3, max_row=3, values_only=True))[0]

        # ตรวจสอบ column B และ C
        col_b = str(header_row[1] or '').strip() if len(header_row) > 1 else ''
        col_c = str(header_row[2] or '').strip() if len(header_row) > 2 else ''

        # ถ้า column B = "ยศ" และ column C = "ชื่อ" → ฟอร์มใหม่
        if col_b == 'ยศ' and col_c == 'ชื่อ':
            return 'NEW'
        # ถ้า column B มีคำว่า "ยศ - ชื่อ" หรือ "ยศ-ชื่อ" → ฟอร์มเก่า
        elif 'ยศ' in col_b and 'ชื่อ' in col_b:
            return 'OLD'
        else:
            # Default เป็นฟอร์มใหม่
            return 'NEW'
    except Exception as e:
        print(f"Warning: Cannot detect format, defaulting to NEW format. Error: {e}")
        return 'NEW'


def split_name_from_old_format(full_name):
    """
    แยกชื่อจากฟอร์มเก่า (รวมกัน 1 คอลัมน์) เป็น 2 บรรทัด

    บรรทัด 1: ยศชื่อ (ไม่เว้นวรรค)
    บรรทัด 2: นามสกุล

    รูปแบบที่รองรับ:
    - "นาย ปัญญา สัจธรรม" → ("นายปัญญา", "สัจธรรม")
    - "นางสาว พรพิไล สุขรัง" → ("นางสาวพรพิไล", "สุขรัง")
    - "สมาชิกตรีสมชาย ผดาวัลย์" → ("สมาชิกตรีสมชาย", "ผดาวัลย์")

    Args:
        full_name (str): ชื่อเต็มที่รวมยศ ชื่อ นามสกุล

    Returns:
        tuple: (first_line, last_line)
    """
    if not full_name:
        return ('', '')

    full_name = full_name.strip()

    # แยกด้วยช่องว่าง
    parts = full_name.split()

    if len(parts) >= 2:
        # เอาคำสุดท้ายเป็นนามสกุล ที่เหลือรวมกันเป็นยศชื่อ (ไม่เว้นวรรค)
        last_line = parts[-1]
        first_line = ''.join(parts[:-1])  # รวมทุกอย่างก่อนนามสกุล ไม่เว้นวรรค
    elif len(parts) == 1:
        # มีแค่คำเดียว ใส่เป็น first_line
        first_line = parts[0]
        last_line = ''
    else:
        # ไม่มีข้อมูล
        first_line = ''
        last_line = ''

    return (first_line, last_line)


def get_excel_sheet_names(file_path):
    """
    อ่านชื่อ sheet ทั้งหมดจากไฟล์ Excel

    Args:
        file_path (str): ที่อยู่ไฟล์ Excel

    Returns:
        list: รายการชื่อ sheet ทั้งหมด
    """
    import openpyxl

    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheet_names = wb.sheetnames
        wb.close()
        return sheet_names
    except Exception as e:
        print(f"Error reading sheet names: {e}")
        return []


def parse_excel_staff(file_path, sheet_name=None):
    """
    Parse Excel file สำหรับ import ข้อมูลบุคลากร
    รองรับทั้งฟอร์มเก่าและฟอร์มใหม่อัตโนมัติ

    === ฟอร์มใหม่ ===
    โครงสร้างไฟล์ Excel:
    - Row 1: หัวข้อหลัก
    - Row 2: พื้นที่
    - Row 3: หัวคอลัมน์
    - Row 4: blank row
    - Row 5+: ข้อมูลจริง

    คอลัมน์:
    A: ลำดับ | B: ยศ | C: ชื่อ | D: นามสกุล | E: บัตรประชาชน | F: หน่วยงาน |
    G: ประเภทบุคคล | H: หน้าที่ | I: อายุ | J: ทะเบียนรถ | K: รูปภาพ/หมายเหตุ

    === ฟอร์มเก่า ===
    โครงสร้างไฟล์ Excel:
    - Row 1: หัวข้อหลัก
    - Row 2: พื้นที่
    - Row 3: หัวคอลัมน์
    - Row 4+: ข้อมูลจริง (ไม่มี blank row)

    คอลัมน์:
    A: ลำดับ | B: ยศ - ชื่อ สกุล | C: บัตรประชาชน | D: หน่วยงาน |
    E: ประเภทบุคคล | F: หน้าที่ | G: อายุ | H: ทะเบียนรถ | I: หมายเหตุ

    Args:
        file_path (str): ที่อยู่ไฟล์ Excel
        sheet_name (str, optional): ชื่อ sheet ที่ต้องการอ่าน ถ้าไม่ระบุจะใช้ active sheet

    Returns:
        list: รายการข้อมูลบุคลากรที่ parse แล้ว
    """
    import openpyxl

    wb = openpyxl.load_workbook(file_path)

    # เลือก sheet ตามที่ระบุ หรือใช้ active sheet
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"Reading sheet: {sheet_name}")
    else:
        ws = wb.active
        print(f"Reading active sheet: {ws.title}")

    # ตรวจสอบฟอร์ม
    format_type = detect_excel_format(ws)
    print(f"Detected Excel format: {format_type}")

    data_list = []

    if format_type == 'NEW':
        # ฟอร์มใหม่: เริ่มอ่านจาก row 5 (ข้าม row 4 ที่เป็น blank)
        for row_num, row in enumerate(ws.iter_rows(min_row=5, values_only=True), start=5):
            # ถ้าแถวว่างเปล่า (ลำดับเป็น None) ให้ข้ามไป
            if row[0] is None:
                continue

            # รวม B (ยศ) + C (ชื่อ) เป็น first_line (ไม่เว้นวรรค)
            title = str(row[1] or '').strip()
            first_name = str(row[2] or '').strip()
            first_line = f"{title}{first_name}"  # ไม่เว้นวรรค
            last_line = str(row[3] or '').strip()  # D: นามสกุล

            # ข้ามแถวว่างหรือไม่สมบูรณ์ (ต้องมีชื่อและนามสกุล)
            if not first_line and not last_line:
                continue

            # แยกข้อมูลตามคอลัมน์ (ฟอร์มใหม่)
            # Clean national_id - เก็บแต่ตัวเลข 13 หลัก
            national_id_raw = str(row[4] or '').strip()
            national_id = ''.join(filter(str.isdigit, national_id_raw))[:13]  # เก็บแต่ตัวเลข สูงสุด 13 หลัก

            staff_data = {
                'row_number': row_num,
                'order': row[0],  # A: ลำดับ
                'first_line': first_line,  # B+C: ยศชื่อ (ไม่เว้นวรรค)
                'last_line': last_line,  # D: นามสกุล
                'national_id': national_id,  # E: บัตรประชาชน (เฉพาะตัวเลข)
                'department_name': str(row[5] or '').strip(),  # F: หน่วยงาน
                'person_type': str(row[6] or '').strip(),  # G: ประเภทบุคคล
                'position': str(row[7] or '').strip(),  # H: หน้าที่
                'age': int(row[8]) if row[8] and str(row[8]).strip() else None,  # I: อายุ
                'vehicle_registration': str(row[9] or '').strip() if len(row) > 9 else '',  # J: ทะเบียนรถ
                'notes': str(row[10] or '').strip() if len(row) > 10 else '',  # K: รูปภาพ/หมายเหตุ
            }

            data_list.append(staff_data)

    else:  # OLD format
        # ฟอร์มเก่า: เริ่มอ่านจาก row 4 (ไม่มี blank row)
        for row_num, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
            # ถ้าแถวว่างเปล่า (ลำดับเป็น None) ให้ข้ามไป
            if row[0] is None:
                continue

            # ข้าม header row (ตรวจสอบว่า column A เป็นตัวเลขหรือไม่)
            # ถ้าไม่ใช่ตัวเลข แสดงว่าเป็น header ให้ข้ามไป
            try:
                order_num = int(row[0])
            except (ValueError, TypeError):
                continue

            # แยกชื่อจากคอลัมน์ B เป็น 2 บรรทัด
            full_name = str(row[1] or '').strip()
            first_line, last_line = split_name_from_old_format(full_name)

            # ข้ามแถวว่างหรือไม่สมบูรณ์ (ต้องมีชื่อหรือนามสกุล)
            if not first_line and not last_line:
                continue

            # แยกข้อมูลตามคอลัมน์ (ฟอร์มเก่า)
            # Clean national_id - เก็บแต่ตัวเลข 13 หลัก
            national_id_raw = str(row[2] or '').strip()
            national_id = ''.join(filter(str.isdigit, national_id_raw))[:13]  # เก็บแต่ตัวเลข สูงสุด 13 หลัก

            staff_data = {
                'row_number': row_num,
                'order': order_num,  # A: ลำดับ
                'first_line': first_line,  # แยกจาก B (ยศชื่อ ไม่เว้นวรรค)
                'last_line': last_line,  # แยกจาก B (นามสกุล)
                'national_id': national_id,  # C: บัตรประชาชน (เฉพาะตัวเลข)
                'department_name': str(row[3] or '').strip(),  # D: หน่วยงาน
                'person_type': str(row[4] or '').strip(),  # E: ประเภทบุคคล
                'position': str(row[5] or '').strip(),  # F: หน้าที่/บทบาท
                'age': int(row[6]) if row[6] and str(row[6]).strip() and str(row[6]).isdigit() else None,  # G: อายุ
                'vehicle_registration': str(row[7] or '').strip() if len(row) > 7 else '',  # H: ทะเบียนรถ
                'notes': str(row[8] or '').strip() if len(row) > 8 else '',  # I: หมายเหตุ
            }

            data_list.append(staff_data)

    return data_list
