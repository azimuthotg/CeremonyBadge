from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.db.models import Count, Q, Sum
from django.utils import timezone
from datetime import timedelta
import json
from apps.registry.models import StaffProfile, BadgeRequest
from apps.badges.models import Badge, BadgeType, PrintLog
from apps.accounts.models import Department
from apps.approvals.models import ApprovalLog

# Create your views here.

@login_required
def dashboard_summary(request):
    """Dashboard สรุปภาพรวมทั้งระบบ (สำหรับ Officer/Admin)"""

    # ตรวจสอบสิทธิ์
    if not request.user.can_manage_all():
        return render(request, '403.html', status=403)

    # สถิติหลัก (เฉพาะหน่วยงานที่เปิดใช้งาน)
    total_staff = StaffProfile.objects.filter(
        department__is_active=True
    ).count()

    total_departments = Department.objects.filter(
        is_active=True
    ).count()

    total_badges = Badge.objects.filter(
        is_active=True,
        staff_profile__department__is_active=True
    ).count()

    printed_badges = Badge.objects.filter(
        is_active=True,
        is_printed=True,
        staff_profile__department__is_active=True
    ).count()

    # ความคืบหน้าแต่ละขั้นตอน (Workflow Progress)
    submitted_count = BadgeRequest.objects.filter(
        status__in=['submitted', 'under_review', 'approved', 'badge_created', 'printed', 'completed'],
        staff_profile__department__is_active=True
    ).count()

    approved_count = BadgeRequest.objects.filter(
        status__in=['approved', 'badge_created', 'printed', 'completed'],
        staff_profile__department__is_active=True
    ).count()

    badge_created_count = BadgeRequest.objects.filter(
        status__in=['badge_created', 'printed', 'completed'],
        staff_profile__department__is_active=True
    ).count()

    completed_count = BadgeRequest.objects.filter(
        status='completed',
        staff_profile__department__is_active=True
    ).count()

    # คำนวณ % ของแต่ละขั้นตอน
    submitted_percent = (submitted_count / total_staff * 100) if total_staff > 0 else 0
    approved_percent = (approved_count / total_staff * 100) if total_staff > 0 else 0
    badge_created_percent = (badge_created_count / total_staff * 100) if total_staff > 0 else 0
    printed_percent = (printed_badges / total_staff * 100) if total_staff > 0 else 0
    completed_percent = (completed_count / total_staff * 100) if total_staff > 0 else 0

    # สถิติบัตรตามประเภท (เฉพาะหน่วยงานที่เปิดใช้งาน) พร้อม %
    badge_stats = Badge.objects.filter(
        is_active=True,
        staff_profile__department__is_active=True
    ).values(
        'badge_type__id',
        'badge_type__name',
        'badge_type__color',
        'badge_type__color_code'
    ).annotate(
        total=Count('id')
    ).order_by('badge_type__name')

    # เพิ่ม % ให้กับแต่ละประเภท
    badge_stats_with_percent = []
    for stat in badge_stats:
        percent = (stat['total'] / total_badges * 100) if total_badges > 0 else 0
        badge_stats_with_percent.append({
            'badge_type__id': stat['badge_type__id'],
            'badge_type__name': stat['badge_type__name'],
            'badge_type__color': stat['badge_type__color'],
            'badge_type__color_code': stat['badge_type__color_code'],
            'total': stat['total'],
            'percent': round(percent, 1)
        })

    # Top 5 หน่วยงานที่มีบุคลากรมากที่สุด
    top_departments = Department.objects.filter(
        is_active=True
    ).annotate(
        staff_count=Count('staff_profiles', filter=Q(staff_profiles__isnull=False))
    ).order_by('-staff_count')[:5]

    # คำนวณ % สำหรับ Top 5
    top_departments_with_percent = []
    for dept in top_departments:
        percent = (dept.staff_count / total_staff * 100) if total_staff > 0 else 0
        top_departments_with_percent.append({
            'department': dept,
            'staff_count': dept.staff_count,
            'percent': round(percent, 1)
        })

    # สรุปสถานะงาน
    pending_status = total_staff - submitted_count  # ยังไม่ส่ง
    in_progress_status = submitted_count - completed_count  # อยู่ระหว่างดำเนินการ

    context = {
        # สถิติหลัก 4 cards
        'total_staff': total_staff,
        'total_departments': total_departments,
        'total_badges': total_badges,
        'printed_badges': printed_badges,

        # ความคืบหน้าแต่ละขั้นตอน
        'submitted_count': submitted_count,
        'submitted_percent': round(submitted_percent, 1),
        'approved_count': approved_count,
        'approved_percent': round(approved_percent, 1),
        'badge_created_count': badge_created_count,
        'badge_created_percent': round(badge_created_percent, 1),
        'printed_percent': round(printed_percent, 1),
        'completed_count': completed_count,
        'completed_percent': round(completed_percent, 1),

        # สถิติบัตรตามประเภท (พร้อม %)
        'badge_stats': badge_stats_with_percent,

        # Top 5 หน่วยงาน
        'top_departments': top_departments_with_percent,

        # สรุปสถานะงาน
        'pending_status': pending_status,
        'in_progress_status': in_progress_status,
        'completed_status': completed_count,
    }

    return render(request, 'reports/dashboard_summary.html', context)


@login_required
def report_by_badge_type(request):
    """รายงานตามประเภทบัตร"""

    # ตรวจสอบสิทธิ์
    if not request.user.can_manage_all():
        return render(request, '403.html', status=403)

    # รับ badge_type_id จาก query string
    badge_type_id = request.GET.get('badge_type')

    # ดึงข้อมูลประเภทบัตรทั้งหมด
    badge_types = BadgeType.objects.filter(is_active=True)

    # ถ้าไม่ได้เลือก ใช้ตัวแรก
    if not badge_type_id and badge_types.exists():
        badge_type_id = str(badge_types.first().id)

    selected_badge_type = None
    badge_data = []

    if badge_type_id:
        try:
            selected_badge_type = BadgeType.objects.get(id=badge_type_id)

            # สถิติประเภทบัตรนี้ (เฉพาะหน่วยงานที่เปิดใช้งาน)
            total_count = Badge.objects.filter(
                badge_type=selected_badge_type,
                is_active=True,
                staff_profile__department__is_active=True
            ).count()

            printed_count = Badge.objects.filter(
                badge_type=selected_badge_type,
                is_active=True,
                is_printed=True,
                staff_profile__department__is_active=True
            ).count()

            not_printed_count = total_count - printed_count

            # แยกตามโซน (เฉพาะหน่วยงานที่เปิดใช้งาน)
            zone_stats = Badge.objects.filter(
                badge_type=selected_badge_type,
                is_active=True,
                staff_profile__department__is_active=True
            ).values(
                'staff_profile__zone__id',
                'staff_profile__zone__code',
                'staff_profile__zone__name'
            ).annotate(
                total=Count('id')
            ).order_by('staff_profile__zone__code')

            # Top 5 หน่วยงาน (เฉพาะหน่วยงานที่เปิดใช้งาน)
            top_departments = Badge.objects.filter(
                badge_type=selected_badge_type,
                is_active=True,
                staff_profile__department__is_active=True
            ).values(
                'staff_profile__department__id',
                'staff_profile__department__name'
            ).annotate(
                total=Count('id')
            ).order_by('-total')[:5]

            # รายการบัตรทั้งหมด (เฉพาะหน่วยงานที่เปิดใช้งาน)
            badges_list = Badge.objects.filter(
                badge_type=selected_badge_type,
                is_active=True,
                staff_profile__department__is_active=True
            ).select_related(
                'staff_profile',
                'staff_profile__department',
                'staff_profile__zone'
            ).order_by('-created_at')

            # กรองตาม department
            department_filter = request.GET.get('department')
            if department_filter:
                badges_list = badges_list.filter(staff_profile__department__id=department_filter)

            # กรองตาม zone
            zone_filter = request.GET.get('zone')
            if zone_filter:
                badges_list = badges_list.filter(staff_profile__zone__id=zone_filter)

            # กรองตามสถานะการพิมพ์
            print_status = request.GET.get('print_status')
            if print_status == 'printed':
                badges_list = badges_list.filter(is_printed=True)
            elif print_status == 'not_printed':
                badges_list = badges_list.filter(is_printed=False)

            badge_data = {
                'total_count': total_count,
                'printed_count': printed_count,
                'not_printed_count': not_printed_count,
                'zone_stats': zone_stats,
                'top_departments': top_departments,
                'badges_list': badges_list[:50],  # แสดง 50 รายการแรก
            }

        except BadgeType.DoesNotExist:
            pass

    # ดึงรายการหน่วยงานและโซนสำหรับ filter
    departments = Department.objects.filter(is_active=True).order_by('name')
    from apps.registry.models import Zone
    zones = Zone.objects.filter(is_active=True).order_by('code')

    context = {
        'badge_types': badge_types,
        'selected_badge_type': selected_badge_type,
        'badge_data': badge_data,
        'departments': departments,
        'zones': zones,
        'department_filter': request.GET.get('department'),
        'zone_filter': request.GET.get('zone'),
        'print_status': request.GET.get('print_status'),
    }

    return render(request, 'reports/report_by_badge_type.html', context)


@login_required
def report_by_department(request):
    """รายงานตามหน่วยงาน"""
    from django.db.models import Case, When, IntegerField

    # ตรวจสอบสิทธิ์
    if not request.user.can_manage_all():
        return render(request, '403.html', status=403)

    # เรียง BadgeType ตามมาตรฐาน: ชมพู → แดง → เหลือง → เขียว
    badge_types_ordered = BadgeType.objects.filter(is_active=True).annotate(
        color_sort=Case(
            When(color='pink', then=1),
            When(color='red', then=2),
            When(color='yellow', then=3),
            When(color='green', then=4),
            output_field=IntegerField(),
        )
    ).order_by('color_sort')

    # สรุปทุกหน่วยงาน
    departments_stats = []

    for dept in Department.objects.filter(is_active=True).order_by('name'):
        staff_count = StaffProfile.objects.filter(
            department=dept
        ).count()

        # นับตามประเภทบัตร (เรียงตามมาตรฐาน)
        badge_counts = {}
        for badge_type in badge_types_ordered:
            count = StaffProfile.objects.filter(
                department=dept,
                badge_type=badge_type
            ).count()
            badge_counts[badge_type.color] = count

        # สถานะ
        approved_count = BadgeRequest.objects.filter(
            staff_profile__department=dept,
            status__in=['approved', 'badge_created', 'printed', 'completed']
        ).count()

        pending_count = BadgeRequest.objects.filter(
            staff_profile__department=dept,
            status__in=['draft', 'photo_uploaded', 'ready_to_submit', 'submitted', 'under_review']
        ).count()

        departments_stats.append({
            'department': dept,
            'total': staff_count,
            'badge_counts': badge_counts,
            'approved': approved_count,
            'pending': pending_count,
        })

    # เรียงตามจำนวนบุคลากร
    departments_stats.sort(key=lambda x: x['total'], reverse=True)

    # Top 5
    top_5_departments = departments_stats[:5]

    context = {
        'departments_stats': departments_stats,
        'top_5_departments': top_5_departments,
        'badge_types': badge_types_ordered,  # ใช้ลำดับที่เรียงแล้ว
    }

    return render(request, 'reports/report_by_department.html', context)


@login_required
def badge_receipt_report_pdf(request, department_id):
    """
    รายงานเซ็นรับบัตรตามหน่วยงาน (PDF)
    แสดงรายชื่อบุคลากรที่สร้างบัตรแล้ว พร้อมช่องเซ็นรับ
    """
    from django.http import HttpResponse
    from django.db.models import Case, When, IntegerField
    from django.conf import settings
    from io import BytesIO
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm, mm
    from reportlab.pdfgen import canvas
    from reportlab.platypus import Table, TableStyle, Paragraph
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import os

    # ตรวจสอบสิทธิ์
    if not request.user.can_manage_all():
        return render(request, '403.html', status=403)

    # ดึงข้อมูลหน่วยงาน
    try:
        department = Department.objects.get(id=department_id, is_active=True)
    except Department.DoesNotExist:
        return HttpResponse('ไม่พบหน่วยงาน', status=404)

    # ลงทะเบียนฟอนต์ไทย
    font_path = os.path.join(settings.BASE_DIR, 'static', 'fonts', 'THSarabunNew.ttf')
    font_bold_path = os.path.join(settings.BASE_DIR, 'static', 'fonts', 'THSarabunNew Bold.ttf')

    try:
        pdfmetrics.registerFont(TTFont('THSarabun', font_path))
        pdfmetrics.registerFont(TTFont('THSarabun-Bold', font_bold_path))
    except:
        # ถ้าไม่มีฟอนต์ใช้ default
        pass

    # เรียงตามลำดับมาตรฐาน: สี (pink → red → yellow → green) → โซน (A → B → C...) → ชื่อ
    color_order = Case(
        When(badge_type__color='pink', then=1),
        When(badge_type__color='red', then=2),
        When(badge_type__color='yellow', then=3),
        When(badge_type__color='green', then=4),
        output_field=IntegerField(),
    )

    staff_with_badges = StaffProfile.objects.filter(
        department=department,
        badge_request__status__in=['approved', 'badge_created', 'printed', 'completed']
    ).select_related(
        'badge_type', 'zone', 'badge_request', 'badge'
    ).annotate(
        color_sort=color_order
    ).order_by(
        'color_sort', 'zone__code', 'first_line', 'last_line'
    ).distinct()

    # นับจำนวนตามประเภทบัตร
    badge_types_ordered = BadgeType.objects.filter(is_active=True).annotate(
        color_sort=Case(
            When(color='pink', then=1),
            When(color='red', then=2),
            When(color='yellow', then=3),
            When(color='green', then=4),
            output_field=IntegerField(),
        )
    ).order_by('color_sort')

    badge_type_counts = {}
    for badge_type in badge_types_ordered:
        count = staff_with_badges.filter(badge_type=badge_type).count()
        badge_type_counts[badge_type.name] = count

    # สร้าง PDF
    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    # หัวกระดาศ
    y_position = height - 2*cm

    pdf.setFont('THSarabun-Bold', 20)
    pdf.drawCentredString(width/2, y_position, 'ใบเซ็นรับบัตร')
    y_position -= 0.7*cm

    pdf.setFont('THSarabun-Bold', 16)
    pdf.drawCentredString(width/2, y_position, f'หน่วยงาน: {department.name}')
    y_position -= 0.6*cm

    pdf.setFont('THSarabun', 14)
    pdf.drawCentredString(width/2, y_position, f'วันที่พิมพ์: {timezone.now().strftime("%d/%m/%Y %H:%M")}')
    y_position -= 0.5*cm

    # สรุปจำนวน
    summary_text = f'จำนวนบัตรทั้งหมด: {staff_with_badges.count()} บัตร'
    if badge_type_counts:
        summary_text += ' ('
        summary_text += ', '.join([f'{name}: {count}' for name, count in badge_type_counts.items() if count > 0])
        summary_text += ')'

    pdf.setFont('THSarabun', 14)
    pdf.drawCentredString(width/2, y_position, summary_text)
    y_position -= 1*cm

    # ตาราง
    data = [['ลำดับ', 'เลขบัตร', 'ชื่อ-นามสกุล', 'ประเภท', 'โซน', 'ลายเซ็น']]

    for idx, staff in enumerate(staff_with_badges, 1):
        badge_number = staff.badge.badge_number if hasattr(staff, 'badge') and staff.badge else '-'
        full_name = f"{staff.first_line} {staff.last_line}"
        badge_type = staff.badge_type.name if staff.badge_type else '-'
        zone = staff.zone.code if staff.zone else '-'

        data.append([str(idx), badge_number, full_name, badge_type, zone, ''])

    # สร้างตาราง
    table = Table(data, colWidths=[1.5*cm, 2.5*cm, 7*cm, 3*cm, 1.5*cm, 3*cm])

    table.setStyle(TableStyle([
        ('FONT', (0, 0), (-1, -1), 'THSarabun', 14),
        ('FONT', (0, 0), (-1, 0), 'THSarabun-Bold', 14),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e0e0e0')),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # ลำดับ
        ('ALIGN', (1, 1), (1, -1), 'CENTER'),  # เลขบัตร
        ('ALIGN', (3, 1), (4, -1), 'CENTER'),  # ประเภท, โซน
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))

    # วาดตาราง
    table_width, table_height = table.wrap(width - 3*cm, height)

    # ถ้าตารางสูงเกินหน้า ให้แบ่งหน้า
    if y_position - table_height < 2*cm:
        pdf.showPage()
        y_position = height - 2*cm

    table.drawOn(pdf, 1.5*cm, y_position - table_height)

    pdf.save()
    buffer.seek(0)

    # Return PDF
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="badge_receipt_{department.name}.pdf"'

    return response



@login_required
def badge_printing_status_pdf(request, department_id):
    """
    PDF รายงานความพร้อมการพิมพ์บัตร - แยกตามสถานะ
    สำหรับวางแผนการพิมพ์ และติดตามปัญหา
    """
    from django.http import HttpResponse
    from weasyprint import HTML, CSS
    from django.template.loader import render_to_string
    from django.db.models import Case, When, IntegerField
    import os

    # ตรวจสอบสิทธิ์
    if not request.user.can_manage_all():
        return render(request, '403.html', status=403)

    # ดึงข้อมูลหน่วยงาน
    try:
        department = Department.objects.get(id=department_id, is_active=True)
    except Department.DoesNotExist:
        return HttpResponse('ไม่พบหน่วยงาน', status=404)

    # เรียงตามลำดับมาตรฐาน: สี (pink → red → yellow → green) → โซน → ชื่อ
    color_order = Case(
        When(badge_type__color='pink', then=1),
        When(badge_type__color='red', then=2),
        When(badge_type__color='yellow', then=3),
        When(badge_type__color='green', then=4),
        output_field=IntegerField(),
    )

    # ดึงบุคลากรทั้งหมดในหน่วยงาน
    all_staff = StaffProfile.objects.filter(
        department=department
    ).select_related(
        'badge_type', 'zone', 'badge_request', 'badge'
    ).annotate(
        color_sort=color_order
    ).order_by(
        'color_sort', 'zone__code', 'first_line', 'last_line'
    ).distinct()

    # ฟังก์ชันช่วยสร้างหมายเหตุ/ปัญหา
    def get_issue_notes(profile):
        """สร้างหมายเหตุอัตโนมัติตามสถานะและข้อมูล"""
        notes = []

        # ตรวจสอบรูปถ่าย (บัตรชมพู/แดง ต้องมีรูป)
        if profile.badge_type and profile.badge_type.color in ['pink', 'red']:
            if not hasattr(profile, 'photo') or not profile.photo:
                notes.append('🚫 ยังไม่มีรูปถ่าย')

        # ตรวจสอบสถานะ
        if hasattr(profile, 'badge_request') and profile.badge_request:
            status = profile.badge_request.status

            if status == 'draft':
                notes.append('📝 ยังไม่ส่งข้อมูล')
            elif status in ['photo_uploaded', 'ready_to_submit']:
                notes.append('🔄 รอส่งข้อมูล')
            elif status == 'rejected':
                # ดึง comment จาก ApprovalLog ล่าสุด
                last_log = profile.badge_request.approval_logs.filter(
                    action='reject'
                ).order_by('-created_at').first()
                if last_log and last_log.comment:
                    notes.append(f'⚠️ {last_log.comment}')
                else:
                    notes.append('⚠️ ส่งกลับแก้ไข')
        else:
            notes.append('📝 ยังไม่เริ่มกรอกข้อมูล')

        return ' | '.join(notes) if notes else '-'

    # แบ่งกลุ่มตามสถานะ
    group_a = []  # พร้อมพิมพ์
    group_b = []  # พิมพ์แล้ว
    group_c = []  # รออนุมัติ/ออกเลข
    group_d = []  # ยังไม่พร้อม

    for staff in all_staff:
        data = {
            'staff': staff,
            'badge_number': staff.badge.badge_number if hasattr(staff, 'badge') and staff.badge else '-',
            'full_name': f"{staff.first_line} {staff.last_line}",
            'badge_type_name': staff.badge_type.name if staff.badge_type else '-',
            'zone_code': staff.zone.code if staff.zone else '-',
            'status': '-',
            'notes': '',
        }

        if hasattr(staff, 'badge_request') and staff.badge_request:
            status = staff.badge_request.status

            # กลุ่ม B: พิมพ์แล้ว
            if status in ['printed', 'completed']:
                data['status'] = 'พิมพ์แล้ว' if status == 'printed' else 'เสร็จสมบูรณ์'
                if hasattr(staff, 'badge') and staff.badge and staff.badge.printed_at:
                    data['printed_date'] = staff.badge.printed_at.strftime('%d/%m/%y')
                else:
                    data['printed_date'] = '-'
                group_b.append(data)

            # กลุ่ม A: พร้อมพิมพ์
            elif status == 'badge_created':
                data['status'] = 'พร้อมพิมพ์'
                group_a.append(data)

            # กลุ่ม C: รออนุมัติ/ออกเลข
            elif status in ['approved', 'submitted', 'under_review']:
                if status == 'approved':
                    data['status'] = 'อนุมัติแล้ว'
                    data['notes'] = 'รอออกเลขบัตร'
                elif status == 'under_review':
                    data['status'] = 'รอตรวจสอบ'
                else:
                    data['status'] = 'ส่งข้อมูลแล้ว'
                group_c.append(data)

            # กลุ่ม D: ยังไม่พร้อม
            else:
                if status == 'draft':
                    data['status'] = 'ร่าง'
                elif status == 'photo_uploaded':
                    data['status'] = 'อัพโหลดรูปแล้ว'
                elif status == 'ready_to_submit':
                    data['status'] = 'พร้อมส่ง'
                elif status == 'rejected':
                    data['status'] = 'ส่งกลับแก้ไข'

                data['notes'] = get_issue_notes(staff)
                group_d.append(data)
        else:
            # ไม่มี badge_request เลย
            data['status'] = 'ยังไม่เริ่ม'
            data['notes'] = get_issue_notes(staff)
            group_d.append(data)

    # คำนวณสถิติ
    total_staff = all_staff.count()
    count_a = len(group_a)
    count_b = len(group_b)
    count_c = len(group_c)
    count_d = len(group_d)

    percent_a = round((count_a / total_staff * 100), 1) if total_staff > 0 else 0
    percent_b = round((count_b / total_staff * 100), 1) if total_staff > 0 else 0
    percent_c = round((count_c / total_staff * 100), 1) if total_staff > 0 else 0
    percent_d = round((count_d / total_staff * 100), 1) if total_staff > 0 else 0

    # สร้าง context สำหรับ template
    context = {
        'department': department,
        'total_staff': total_staff,
        'group_a': group_a,
        'group_b': group_b,
        'group_c': group_c,
        'group_d': group_d,
        'count_a': count_a,
        'count_b': count_b,
        'count_c': count_c,
        'count_d': count_d,
        'percent_a': percent_a,
        'percent_b': percent_b,
        'percent_c': percent_c,
        'percent_d': percent_d,
        'generated_date': timezone.now(),
    }

    # Render HTML template
    html_string = render_to_string('reports/badge_printing_status_pdf.html', context)

    # สร้าง PDF
    from django.conf import settings
    from weasyprint.text.fonts import FontConfiguration

    # Font configuration for Thai fonts
    font_config = FontConfiguration()

    # Path to Thai font
    font_path = os.path.join(settings.BASE_DIR, 'static', 'fonts', 'THSarabunNew.ttf')

    # CSS with embedded font
    css_string = f'''
        @font-face {{
            font-family: 'THSarabunNew';
            src: url('file://{font_path}') format('truetype');
        }}
        @page {{
            size: A4;
            margin: 1.5cm 1cm;
        }}
        * {{
            font-family: 'THSarabunNew', 'DejaVu Sans', sans-serif;
        }}
        body {{
            font-size: 14pt;
            line-height: 1.4;
        }}
        .header {{
            text-align: center;
            margin-bottom: 0.5cm;
        }}
        h1 {{
            font-size: 18pt;
            font-weight: bold;
            margin: 0;
        }}
        h2 {{
            font-size: 16pt;
            font-weight: bold;
            margin-top: 0.5cm;
            margin-bottom: 0.2cm;
            padding: 5px;
            background-color: #f0f0f0;
        }}
        .summary {{
            border: 2px solid #333;
            padding: 10px;
            margin-bottom: 0.5cm;
            background-color: #f9f9f9;
        }}
        .summary-row {{
            display: flex;
            justify-content: space-between;
            margin: 3px 0;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            margin-bottom: 0.3cm;
            font-size: 13pt;
        }}
        th, td {{
            border: 1px solid #000;
            padding: 5px 4px;
            text-align: left;
        }}
        th {{
            background-color: #e0e0e0;
            font-weight: bold;
            text-align: center;
        }}
        .text-center {{
            text-align: center;
        }}
        .checkbox {{
            width: 15px;
            height: 15px;
            border: 1px solid #000;
            display: inline-block;
        }}
        .group-a {{ background-color: #d4edda; }}
        .group-b {{ background-color: #d1ecf1; }}
        .group-c {{ background-color: #fff3cd; }}
        .group-d {{ background-color: #f8d7da; }}
    '''

    html = HTML(string=html_string, base_url=request.build_absolute_uri())
    pdf_file = html.write_pdf(
        stylesheets=[CSS(string=css_string, font_config=font_config)],
        font_config=font_config
    )

    # Return PDF
    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="printing_status_{department.name}.pdf"'

    return response


@login_required


@login_required
def print_manager_dashboard(request):
    """
    หน้าเว็บ Print Manager - จัดการและติดตามการพิมพ์บัตร
    แสดงภาพรวมทุกหน่วยงาน พร้อมฟังก์ชันกรองและ export PDF
    """
    from django.db.models import Case, When, IntegerField, Q

    # ตรวจสอบสิทธิ์
    if not request.user.can_manage_all():
        return render(request, '403.html', status=403)

    # ดึงพารามิเตอร์จาก query string
    department_filter = request.GET.get('department')
    status_filter = request.GET.get('status')  # ready, printed, waiting, not_ready
    search_query = request.GET.get('search', '').strip()

    # เรียงตามลำดับมาตรฐาน: สี → โซน → ชื่อ
    color_order = Case(
        When(badge_type__color='pink', then=1),
        When(badge_type__color='red', then=2),
        When(badge_type__color='yellow', then=3),
        When(badge_type__color='green', then=4),
        output_field=IntegerField(),
    )

    # สถิติแต่ละหน่วยงาน
    departments_stats = []

    for dept in Department.objects.filter(is_active=True).order_by('name'):
        total_count = StaffProfile.objects.filter(department=dept).count()

        if total_count == 0:
            continue

        # นับแต่ละกลุ่ม
        ready_count = BadgeRequest.objects.filter(
            staff_profile__department=dept,
            status='badge_created'
        ).count()

        printed_count = BadgeRequest.objects.filter(
            staff_profile__department=dept,
            status__in=['printed', 'completed']
        ).count()

        waiting_count = BadgeRequest.objects.filter(
            staff_profile__department=dept,
            status__in=['approved', 'submitted', 'under_review']
        ).count()

        not_ready_count = total_count - ready_count - printed_count - waiting_count

        # คำนวณ %
        ready_percent = round((ready_count / total_count * 100), 1) if total_count > 0 else 0
        printed_percent = round((printed_count / total_count * 100), 1) if total_count > 0 else 0

        departments_stats.append({
            'department': dept,
            'total': total_count,
            'ready': ready_count,
            'printed': printed_count,
            'waiting': waiting_count,
            'not_ready': not_ready_count,
            'ready_percent': ready_percent,
            'printed_percent': printed_percent,
        })

    # กรองตามหน่วยงาน
    if department_filter:
        departments_stats = [d for d in departments_stats if str(d['department'].id) == department_filter]

    # ถ้าเลือกดูรายละเอียดหน่วยงาน
    selected_department = None
    staff_data = []

    if department_filter:
        try:
            selected_department = Department.objects.get(id=department_filter, is_active=True)

            # ดึงบุคลากรทั้งหมด
            all_staff = StaffProfile.objects.filter(
                department=selected_department
            ).select_related(
                'badge_type', 'zone', 'badge_request', 'badge'
            ).annotate(
                color_sort=color_order
            ).order_by(
                'color_sort', 'zone__code', 'first_line', 'last_line'
            ).distinct()

            # ค้นหา
            if search_query:
                all_staff = all_staff.filter(
                    Q(first_line__icontains=search_query) |
                    Q(last_line__icontains=search_query) |
                    Q(badge__badge_number__icontains=search_query)
                )

            # แบ่งกลุ่มและกรองตามสถานะ
            for staff in all_staff:
                # กำหนดกลุ่ม
                group = 'not_ready'
                status_text = 'ยังไม่เริ่ม'
                badge_class = 'danger'

                if hasattr(staff, 'badge_request') and staff.badge_request:
                    status = staff.badge_request.status

                    if status in ['printed', 'completed']:
                        group = 'printed'
                        status_text = 'พิมพ์แล้ว' if status == 'printed' else 'เสร็จสมบูรณ์'
                        badge_class = 'info'
                    elif status == 'badge_created':
                        group = 'ready'
                        status_text = 'พร้อมพิมพ์'
                        badge_class = 'success'
                    elif status in ['approved', 'submitted', 'under_review']:
                        group = 'waiting'
                        if status == 'approved':
                            status_text = 'รอออกเลข'
                        elif status == 'under_review':
                            status_text = 'รอตรวจสอบ'
                        else:
                            status_text = 'ส่งแล้ว'
                        badge_class = 'warning'
                    else:
                        if status == 'draft':
                            status_text = 'ร่าง'
                        elif status == 'photo_uploaded':
                            status_text = 'อัพโหลดรูปแล้ว'
                        elif status == 'ready_to_submit':
                            status_text = 'พร้อมส่ง'
                        elif status == 'rejected':
                            status_text = 'ส่งกลับแก้ไข'

                # กรองตามสถานะที่เลือก
                if status_filter and group != status_filter:
                    continue

                # สร้าง notes
                notes = []
                if staff.badge_type and staff.badge_type.color in ['pink', 'red']:
                    if not hasattr(staff, 'photo') or not staff.photo:
                        notes.append('ไม่มีรูป')

                staff_data.append({
                    'id': staff.id,
                    'badge_number': staff.badge.badge_number if hasattr(staff, 'badge') and staff.badge else '-',
                    'full_name': f"{staff.first_line} {staff.last_line}",
                    'badge_type': staff.badge_type.name if staff.badge_type else '-',
                    'badge_color': staff.badge_type.color if staff.badge_type else '',
                    'zone': staff.zone.code if staff.zone else '-',
                    'status': status_text,
                    'group': group,
                    'badge_class': badge_class,
                    'notes': ', '.join(notes) if notes else '',
                })

        except Department.DoesNotExist:
            pass

    # รายการหน่วยงานสำหรับ dropdown
    departments = Department.objects.filter(is_active=True).order_by('name')

    context = {
        'departments_stats': departments_stats,
        'departments': departments,
        'selected_department': selected_department,
        'staff_data': staff_data,
        'department_filter': department_filter,
        'status_filter': status_filter,
        'search_query': search_query,
    }

    return render(request, 'reports/print_manager.html', context)


@login_required
def submitter_report(request):
    """รายงานหน่วยงาน (สำหรับ Submitter)"""
    from django.db.models import Case, When, IntegerField

    # Submitter ดูได้เฉพาะหน่วยงานของตัวเอง
    if request.user.role != 'submitter':
        return render(request, '403.html', status=403)

    department = request.user.department

    # เรียง BadgeType ตามมาตรฐาน: ชมพู → แดง → เหลือง → เขียว
    badge_types_ordered = BadgeType.objects.filter(is_active=True).annotate(
        color_sort=Case(
            When(color='pink', then=1),
            When(color='red', then=2),
            When(color='yellow', then=3),
            When(color='green', then=4),
            output_field=IntegerField(),
        )
    ).order_by('color_sort')

    # สถิติรวม
    total_staff = StaffProfile.objects.filter(
        department=department
    ).count()

    pending_requests = BadgeRequest.objects.filter(
        staff_profile__department=department,
        status__in=['draft', 'photo_uploaded', 'ready_to_submit']
    ).count()

    submitted_requests = BadgeRequest.objects.filter(
        staff_profile__department=department,
        status__in=['submitted', 'under_review']
    ).count()

    approved_badges = BadgeRequest.objects.filter(
        staff_profile__department=department,
        status__in=['approved', 'badge_created', 'printed', 'completed']
    ).count()

    rejected_requests = BadgeRequest.objects.filter(
        staff_profile__department=department,
        status='rejected'
    ).count()

    # แยกตามประเภทบัตร
    badge_type_stats = StaffProfile.objects.filter(
        department=department
    ).values(
        'badge_type__id',
        'badge_type__name',
        'badge_type__color',
        'badge_type__color_code'
    ).annotate(
        total=Count('id')
    ).order_by('badge_type__name')

    # รายการบุคลากร
    staff_list = StaffProfile.objects.filter(
        department=department
    ).select_related(
        'badge_type',
        'zone'
    ).prefetch_related(
        'badge_requests'
    ).order_by('-created_at')

    # กรองตามสถานะ
    status_filter = request.GET.get('status')
    if status_filter:
        staff_list = staff_list.filter(
            badge_requests__status=status_filter
        )

    # กรองตามประเภทบัตร
    badge_type_filter = request.GET.get('badge_type')
    if badge_type_filter:
        staff_list = staff_list.filter(badge_type__id=badge_type_filter)

    context = {
        'department': department,
        'total_staff': total_staff,
        'pending_requests': pending_requests,
        'submitted_requests': submitted_requests,
        'approved_badges': approved_badges,
        'rejected_requests': rejected_requests,
        'badge_type_stats': badge_type_stats,
        'staff_list': staff_list,
        'badge_types': badge_types_ordered,  # ใช้ลำดับที่เรียงแล้ว
        'status_filter': status_filter,
        'badge_type_filter': badge_type_filter,
    }

    return render(request, 'reports/submitter_report.html', context)


def public_status_dashboard(request):
    """
    Dashboard สาธารณะ - แสดงสถานะการส่งข้อมูลและความคืบหน้าของแต่ละหน่วยงาน
    ไม่ต้อง login, ไม่แสดงข้อมูลส่วนบุคคล (Privacy-first design)
    """

    # สถิติรวมทั้งระบบ
    total_departments = Department.objects.filter(is_active=True).count()
    total_staff = StaffProfile.objects.count()

    # นับสถานะต่างๆ
    total_submitted = BadgeRequest.objects.filter(
        status__in=['submitted', 'under_review', 'approved', 'badge_created', 'printed', 'completed']
    ).count()

    total_approved = BadgeRequest.objects.filter(
        status__in=['approved', 'badge_created', 'printed', 'completed']
    ).count()

    total_badges = Badge.objects.filter(is_active=True).count()

    total_printed = Badge.objects.filter(
        is_active=True,
        is_printed=True
    ).count()

    # คำนวณ % ความคืบหน้าโดยรวม
    overall_progress = round((total_approved / total_staff * 100), 1) if total_staff > 0 else 0

    # สถิติแยกตามหน่วยงาน
    department_stats = []

    for dept in Department.objects.filter(is_active=True).order_by('order', 'name'):
        total_dept_staff = StaffProfile.objects.filter(department=dept).count()

        if total_dept_staff == 0:
            continue  # ข้ามหน่วยงานที่ยังไม่มีข้อมูล

        # นับสถานะต่างๆ ของหน่วยงาน
        submitted_count = BadgeRequest.objects.filter(
            staff_profile__department=dept,
            status__in=['submitted', 'under_review', 'approved', 'badge_created', 'printed', 'completed']
        ).count()

        approved_count = BadgeRequest.objects.filter(
            staff_profile__department=dept,
            status__in=['approved', 'badge_created', 'printed', 'completed']
        ).count()

        printed_count = Badge.objects.filter(
            staff_profile__department=dept,
            is_active=True,
            is_printed=True
        ).count()

        # คำนวณเปอร์เซ็นต์
        submit_percentage = round((submitted_count / total_dept_staff * 100), 1) if total_dept_staff > 0 else 0
        approve_percentage = round((approved_count / total_dept_staff * 100), 1) if total_dept_staff > 0 else 0
        print_percentage = round((printed_count / total_dept_staff * 100), 1) if total_dept_staff > 0 else 0

        # กำหนดสถานะ badge
        if submit_percentage >= 100:
            status_badge = 'success'  # เขียว - ส่งครบ
            status_text = 'ส่งครบแล้ว'
        elif submit_percentage >= 50:
            status_badge = 'warning'  # เหลือง - ส่งบางส่วน
            status_text = 'ส่งบางส่วน'
        elif submit_percentage > 0:
            status_badge = 'info'  # ฟ้า - เริ่มส่งแล้ว
            status_text = 'กำลังดำเนินการ'
        else:
            status_badge = 'danger'  # แดง - ยังไม่ส่ง
            status_text = 'ยังไม่ส่งข้อมูล'

        department_stats.append({
            'department': dept,
            'total': total_dept_staff,
            'submitted': submitted_count,
            'approved': approved_count,
            'printed': printed_count,
            'submit_percentage': submit_percentage,
            'approve_percentage': approve_percentage,
            'print_percentage': print_percentage,
            'status_badge': status_badge,
            'status_text': status_text,
        })

    # เรียงลำดับตาม percentage (มากไปน้อย)
    department_stats.sort(key=lambda x: x['submit_percentage'], reverse=True)

    # สถิติแยกตามประเภทบัตร (Badge Type)
    badge_type_stats = []

    for badge_type in BadgeType.objects.filter(is_active=True).order_by('name'):
        total_type = StaffProfile.objects.filter(badge_type=badge_type).count()

        approved_type = BadgeRequest.objects.filter(
            staff_profile__badge_type=badge_type,
            status__in=['approved', 'badge_created', 'printed', 'completed']
        ).count()

        printed_type = Badge.objects.filter(
            badge_type=badge_type,
            is_active=True,
            is_printed=True
        ).count()

        approve_percentage = round((approved_type / total_type * 100), 1) if total_type > 0 else 0
        print_percentage = round((printed_type / total_type * 100), 1) if total_type > 0 else 0

        badge_type_stats.append({
            'badge_type': badge_type,
            'total': total_type,
            'approved': approved_type,
            'printed': printed_type,
            'approve_percentage': approve_percentage,
            'print_percentage': print_percentage,
        })

    # สถิติสถานะ (Status Distribution) สำหรับ Pie Chart
    status_distribution = []

    status_groups = [
        {
            'label': 'ยังไม่ส่ง',
            'statuses': ['draft', 'photo_uploaded', 'ready_to_submit'],
            'color': '#EF4444'  # Red
        },
        {
            'label': 'รอตรวจสอบ',
            'statuses': ['submitted', 'under_review'],
            'color': '#F59E0B'  # Orange
        },
        {
            'label': 'อนุมัติแล้ว',
            'statuses': ['approved', 'badge_created'],
            'color': '#3B82F6'  # Blue
        },
        {
            'label': 'พิมพ์แล้ว',
            'statuses': ['printed', 'completed'],
            'color': '#10B981'  # Green
        },
        {
            'label': 'ส่งกลับแก้ไข',
            'statuses': ['rejected'],
            'color': '#8B5CF6'  # Purple
        },
    ]

    for group in status_groups:
        count = BadgeRequest.objects.filter(status__in=group['statuses']).count()
        percentage = round((count / total_staff * 100), 1) if total_staff > 0 else 0

        status_distribution.append({
            'label': group['label'],
            'count': count,
            'percentage': percentage,
            'color': group['color'],
        })

    # แปลงเป็น JSON สำหรับ Chart.js
    status_labels = [item['label'] for item in status_distribution]
    status_counts = [item['count'] for item in status_distribution]
    status_colors = [item['color'] for item in status_distribution]

    # เวลาอัพเดทล่าสุด
    last_updated = timezone.now()

    context = {
        # สถิติหลัก
        'total_departments': total_departments,
        'total_staff': total_staff,
        'total_submitted': total_submitted,
        'total_approved': total_approved,
        'total_badges': total_badges,
        'total_printed': total_printed,
        'overall_progress': overall_progress,

        # สถิติหน่วยงาน
        'department_stats': department_stats,

        # สถิติประเภทบัตร
        'badge_type_stats': badge_type_stats,

        # สถิติสถานะ (สำหรับ chart)
        'status_distribution': status_distribution,
        'status_labels': json.dumps(status_labels),
        'status_counts': json.dumps(status_counts),
        'status_colors': json.dumps(status_colors),

        # ข้อมูลเพิ่มเติม
        'last_updated': last_updated,
    }

    return render(request, 'public/status_dashboard.html', context)


@login_required
def duplicate_check_view(request):
    """
    ตรวจสอบข้อมูลซ้ำ - หน้าตรวจสอบชื่อซ้ำ, บัตรประชาชนซ้ำ, และไม่มีบัตรประชาชน
    สำหรับ Officer/Admin เท่านั้น
    """
    from django.db.models import Case, When, IntegerField

    # ตรวจสอบสิทธิ์
    if not request.user.can_manage_all():
        return render(request, '403.html', status=403)

    # สร้างลำดับการเรียงตามสีบัตร (มาตรฐาน: ชมพู → แดง → เหลือง → เขียว)
    color_order = Case(
        When(badge_type__color='pink', then=1),
        When(badge_type__color='red', then=2),
        When(badge_type__color='yellow', then=3),
        When(badge_type__color='green', then=4),
        output_field=IntegerField(),
    )

    # =====================
    # Tab 1: ตรวจสอบชื่อซ้ำ
    # =====================

    # หาชื่อที่ซ้ำ (first_line + last_line เหมือนกัน) - เฉพาะหน่วยงานที่เปิดใช้งาน
    name_duplicates = StaffProfile.objects.filter(
        department__is_active=True
    ).values(
        'first_line',
        'last_line'
    ).annotate(
        count=Count('id')
    ).filter(
        count__gt=1
    ).order_by('-count', 'first_line', 'last_line')

    # สร้างรายละเอียดแต่ละกลุ่มชื่อซ้ำ
    name_duplicate_groups = []

    for item in name_duplicates:
        # ดึงรายการทั้งหมดที่มีชื่อนี้ - เฉพาะหน่วยงานที่เปิดใช้งาน
        # เรียงตาม: หน่วยงาน → สีบัตร → โซน → ชื่อ
        profiles = StaffProfile.objects.filter(
            first_line=item['first_line'],
            last_line=item['last_line'],
            department__is_active=True
        ).select_related(
            'department',
            'badge_type',
            'zone'
        ).prefetch_related('badge_request').annotate(
            color_sort=color_order
        ).order_by(
            'department__name', 'color_sort', 'zone__code'
        )

        # รวบรวมข้อมูลหน่วยงาน
        departments = set()
        badge_types = set()

        for profile in profiles:
            departments.add(profile.department.name if profile.department else '-')
            badge_types.add(profile.badge_type.name if profile.badge_type else '-')

        name_duplicate_groups.append({
            'full_name': f"{item['first_line']} {item['last_line']}",
            'first_line': item['first_line'],
            'last_line': item['last_line'],
            'count': item['count'],
            'profiles': profiles,
            'departments': ', '.join(sorted(departments)),
            'badge_types': ', '.join(sorted(badge_types)),
        })

    # ====================================
    # Tab 2: ตรวจสอบบัตรประชาชนซ้ำ
    # ====================================

    # หาบัตรประชาชนที่ซ้ำ (ไม่เอา NULL) - เฉพาะหน่วยงานที่เปิดใช้งาน
    national_id_duplicates = StaffProfile.objects.filter(
        national_id__isnull=False,
        department__is_active=True
    ).exclude(
        national_id=''
    ).values(
        'national_id'
    ).annotate(
        count=Count('id')
    ).filter(
        count__gt=1
    ).order_by('-count', 'national_id')

    # สร้างรายละเอียดแต่ละกลุ่มบัตรประชาชนซ้ำ
    national_id_duplicate_groups = []

    for item in national_id_duplicates:
        # ดึงรายการทั้งหมดที่มีบัตรประชาชนนี้ - เฉพาะหน่วยงานที่เปิดใช้งาน
        # เรียงตาม: หน่วยงาน → สีบัตร → โซน → ชื่อ
        profiles = StaffProfile.objects.filter(
            national_id=item['national_id'],
            department__is_active=True
        ).select_related(
            'department',
            'badge_type',
            'zone'
        ).prefetch_related('badge_request').annotate(
            color_sort=color_order
        ).order_by(
            'department__name', 'color_sort', 'zone__code', 'first_line', 'last_line'
        )

        national_id_duplicate_groups.append({
            'national_id': item['national_id'],
            'count': item['count'],
            'profiles': profiles,
        })

    # ==========================================
    # Tab 3: รายการไม่มีบัตรประชาชน
    # ==========================================

    # หารายการที่ไม่มีบัตรประชาชน - เฉพาะหน่วยงานที่เปิดใช้งาน
    # เรียงตาม: หน่วยงาน → สีบัตร → โซน → ชื่อ
    no_national_id_profiles = StaffProfile.objects.filter(
        Q(national_id__isnull=True) | Q(national_id=''),
        department__is_active=True
    ).select_related(
        'department',
        'badge_type',
        'zone'
    ).prefetch_related('badge_request').annotate(
        color_sort=color_order
    ).order_by(
        'department__name', 'color_sort', 'zone__code', 'first_line', 'last_line'
    )

    # ฟิลเตอร์ตามหน่วยงาน (ถ้ามี)
    department_filter = request.GET.get('department')
    if department_filter:
        no_national_id_profiles = no_national_id_profiles.filter(
            department__id=department_filter
        )

    # ==========================================
    # Tab 4: ไม่มีรูปถ่าย (เฉพาะบัตรชมพู/แดง)
    # ==========================================

    # หารายการบัตรชมพู/แดงที่ไม่มีรูปถ่าย - เฉพาะหน่วยงานที่เปิดใช้งาน
    # เรียงตาม: หน่วยงาน → สีบัตร → โซน → ชื่อ
    no_photo_profiles = StaffProfile.objects.filter(
        badge_type__color__in=['pink', 'red'],
        department__is_active=True
    ).exclude(
        id__in=StaffProfile.objects.filter(photo__isnull=False).values('id')
    ).select_related(
        'department',
        'badge_type',
        'zone'
    ).prefetch_related('badge_request').annotate(
        color_sort=color_order
    ).order_by(
        'department__name', 'color_sort', 'zone__code', 'first_line', 'last_line'
    )

    # ฟิลเตอร์ตามหน่วยงาน (ถ้ามี Tab 4)
    department_filter_photo = request.GET.get('department_photo')
    if department_filter_photo:
        no_photo_profiles = no_photo_profiles.filter(
            department__id=department_filter_photo
        )

    # สถิติรวม (เฉพาะหน่วยงานที่เปิดใช้งาน)
    total_staff = StaffProfile.objects.filter(department__is_active=True).count()
    total_name_duplicates = sum([item['count'] for item in name_duplicates])
    total_national_id_duplicates = sum([item['count'] for item in national_id_duplicates])
    total_no_national_id = no_national_id_profiles.count()
    total_no_photo = no_photo_profiles.count()

    # สถิติเพิ่มเติม
    unique_name_duplicate_groups = len(name_duplicate_groups)
    unique_national_id_duplicate_groups = len(national_id_duplicate_groups)

    # รายการหน่วยงานสำหรับ filter
    departments = Department.objects.filter(is_active=True).order_by('name')

    context = {
        # Tab 1: ชื่อซ้ำ
        'name_duplicate_groups': name_duplicate_groups,
        'unique_name_duplicate_groups': unique_name_duplicate_groups,
        'total_name_duplicates': total_name_duplicates,

        # Tab 2: บัตรประชาชนซ้ำ
        'national_id_duplicate_groups': national_id_duplicate_groups,
        'unique_national_id_duplicate_groups': unique_national_id_duplicate_groups,
        'total_national_id_duplicates': total_national_id_duplicates,

        # Tab 3: ไม่มีบัตรประชาชน
        'no_national_id_profiles': no_national_id_profiles,
        'total_no_national_id': total_no_national_id,

        # Tab 4: ไม่มีรูปถ่าย
        'no_photo_profiles': no_photo_profiles,
        'total_no_photo': total_no_photo,

        # สถิติรวม
        'total_staff': total_staff,

        # ฟิลเตอร์
        'departments': departments,
        'department_filter': department_filter,
        'department_filter_photo': department_filter_photo,
    }

    return render(request, 'reports/duplicate_check.html', context)


@login_required
def department_detailed_report_pdf(request, department_id):
    """
    รายงาน PDF รายละเอียดหน่วยงาน
    แสดงสถิติบัตรแต่ละสี และรายการบัตรทั้งหมดเรียงตามเลขบัตร
    """
    from django.http import HttpResponse
    from weasyprint import HTML, CSS
    from django.template.loader import render_to_string
    from django.db.models import Case, When, IntegerField, Count
    from django.conf import settings
    import os

    # ตรวจสอบสิทธิ์
    if not request.user.can_manage_all():
        return render(request, '403.html', status=403)

    # ดึงข้อมูลหน่วยงาน
    try:
        department = Department.objects.get(id=department_id, is_active=True)
    except Department.DoesNotExist:
        return HttpResponse('ไม่พบหน่วยงาน', status=404)

    # เรียง BadgeType ตามมาตรฐาน: ชมพู → แดง → เหลือง → เขียว
    badge_types_ordered = BadgeType.objects.filter(is_active=True).annotate(
        color_sort=Case(
            When(color='pink', then=1),
            When(color='red', then=2),
            When(color='yellow', then=3),
            When(color='green', then=4),
            output_field=IntegerField(),
        )
    ).order_by('color_sort')

    # นับจำนวนแต่ละประเภทบัตร
    badge_type_stats = []
    total_badges = 0

    for badge_type in badge_types_ordered:
        count = Badge.objects.filter(
            staff_profile__department=department,
            badge_type=badge_type,
            is_active=True
        ).count()

        printed_count = Badge.objects.filter(
            staff_profile__department=department,
            badge_type=badge_type,
            is_active=True,
            is_printed=True
        ).count()

        if count > 0:
            badge_type_stats.append({
                'badge_type': badge_type,
                'count': count,
                'printed_count': printed_count,
                'not_printed_count': count - printed_count,
            })
            total_badges += count

    # ดึงรายการบัตรทั้งหมดเรียงตามเลขบัตร
    badges_list = Badge.objects.filter(
        staff_profile__department=department,
        is_active=True
    ).select_related(
        'staff_profile',
        'badge_type',
        'staff_profile__zone'
    ).order_by('badge_number')

    # สร้าง context สำหรับ template
    context = {
        'department': department,
        'badge_type_stats': badge_type_stats,
        'total_badges': total_badges,
        'badges_list': badges_list,
        'generated_date': timezone.now(),
    }

    # Render HTML template
    html_string = render_to_string('reports/department_detailed_report_pdf.html', context)

    # สร้าง PDF
    from weasyprint.text.fonts import FontConfiguration

    # Font configuration for Thai fonts
    font_config = FontConfiguration()

    # Path to Thai font
    font_path = os.path.join(settings.BASE_DIR, 'static', 'fonts', 'THSarabunNew.ttf')
    font_bold_path = os.path.join(settings.BASE_DIR, 'static', 'fonts', 'THSarabunNew Bold.ttf')

    # CSS with embedded font
    css_string = f'''
        @font-face {{
            font-family: 'THSarabunNew';
            src: url('file://{font_path}') format('truetype');
            font-weight: normal;
        }}
        @font-face {{
            font-family: 'THSarabunNew';
            src: url('file://{font_bold_path}') format('truetype');
            font-weight: bold;
        }}
        @page {{
            size: A4;
            margin: 1.5cm 1cm;
        }}
        * {{
            font-family: 'THSarabunNew', 'DejaVu Sans', sans-serif;
        }}
        body {{
            font-size: 14pt;
            line-height: 1.4;
        }}
        .header {{
            text-align: center;
            margin-bottom: 0.5cm;
            border-bottom: 2px solid #333;
            padding-bottom: 0.3cm;
        }}
        h1 {{
            font-size: 20pt;
            font-weight: bold;
            margin: 0.2cm 0;
        }}
        h2 {{
            font-size: 16pt;
            font-weight: bold;
            margin: 0.2cm 0;
        }}
        .meta {{
            font-size: 12pt;
            color: #666;
            margin-top: 0.2cm;
        }}
        .summary {{
            margin: 0.5cm 0;
            padding: 0.3cm;
            background-color: #f0f0f0;
            border: 1px solid #ccc;
        }}
        .summary-grid {{
            display: grid;
            grid-template-columns: repeat(4, 1fr);
            gap: 0.3cm;
            margin-top: 0.3cm;
        }}
        .summary-item {{
            text-align: center;
            padding: 0.3cm;
            background-color: white;
            border-radius: 4px;
            border: 2px solid;
        }}
        .summary-item h3 {{
            font-size: 24pt;
            font-weight: bold;
            margin: 0.1cm 0;
        }}
        .summary-item p {{
            font-size: 12pt;
            margin: 0;
            color: #666;
        }}
        .badge-pink {{ border-color: #FFC0CB; background-color: #FFF0F5; }}
        .badge-red {{ border-color: #FF6B6B; background-color: #FFE5E5; }}
        .badge-yellow {{ border-color: #FFD93D; background-color: #FFFBEA; }}
        .badge-green {{ border-color: #6BCF7F; background-color: #E8F8EA; }}

        table {{
            width: 100%;
            border-collapse: collapse;
            margin-top: 0.5cm;
            font-size: 13pt;
        }}
        th, td {{
            border: 1px solid #000;
            padding: 0.15cm 0.2cm;
            text-align: left;
        }}
        th {{
            background-color: #e0e0e0;
            font-weight: bold;
            text-align: center;
        }}
        .text-center {{
            text-align: center;
        }}
        .status-printed {{
            color: #10B981;
            font-weight: bold;
        }}
        .status-not-printed {{
            color: #F59E0B;
        }}
    '''

    html = HTML(string=html_string, base_url=request.build_absolute_uri())
    pdf_file = html.write_pdf(
        stylesheets=[CSS(string=css_string, font_config=font_config)],
        font_config=font_config
    )

    # Return PDF
    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="department_report_{department.name}.pdf"'

    return response


@login_required
def department_badge_type_report_pdf(request, department_id, badge_type_id):
    """
    รายงาน PDF แยกตามสีบัตรของแต่ละหน่วยงาน
    เช่น บัตรแดงของมหาวิทยาลัยนครพนม
    """
    from django.http import HttpResponse
    from weasyprint import HTML, CSS
    from django.template.loader import render_to_string
    from django.conf import settings
    import os

    # ตรวจสอบสิทธิ์
    if not request.user.can_manage_all():
        return render(request, '403.html', status=403)

    # ดึงข้อมูลหน่วยงาน
    try:
        department = Department.objects.get(id=department_id, is_active=True)
    except Department.DoesNotExist:
        return HttpResponse('ไม่พบหน่วยงาน', status=404)

    # ดึงข้อมูลประเภทบัตร
    try:
        badge_type = BadgeType.objects.get(id=badge_type_id, is_active=True)
    except BadgeType.DoesNotExist:
        return HttpResponse('ไม่พบประเภทบัตร', status=404)

    # นับจำนวนบัตรประเภทนี้
    total_count = Badge.objects.filter(
        staff_profile__department=department,
        badge_type=badge_type,
        is_active=True
    ).count()

    printed_count = Badge.objects.filter(
        staff_profile__department=department,
        badge_type=badge_type,
        is_active=True,
        is_printed=True
    ).count()

    not_printed_count = total_count - printed_count

    # ดึงรายการบัตรเรียงตามเลขบัตร
    badges_list = Badge.objects.filter(
        staff_profile__department=department,
        badge_type=badge_type,
        is_active=True
    ).select_related(
        'staff_profile',
        'staff_profile__zone'
    ).order_by('badge_number')

    # สร้าง context
    context = {
        'department': department,
        'badge_type': badge_type,
        'total_count': total_count,
        'printed_count': printed_count,
        'not_printed_count': not_printed_count,
        'badges_list': badges_list,
        'generated_date': timezone.now(),
    }

    # Render HTML template
    html_string = render_to_string('reports/department_badge_type_report_pdf.html', context)

    # สร้าง PDF
    from weasyprint.text.fonts import FontConfiguration

    font_config = FontConfiguration()
    font_path = os.path.join(settings.BASE_DIR, 'static', 'fonts', 'THSarabunNew.ttf')
    font_bold_path = os.path.join(settings.BASE_DIR, 'static', 'fonts', 'THSarabunNew Bold.ttf')

    # กำหนดสีตามประเภทบัตร
    color_map = {
        'pink': '#FFC0CB',
        'red': '#FF6B6B',
        'yellow': '#FFD93D',
        'green': '#6BCF7F'
    }
    badge_color = color_map.get(badge_type.color, '#999999')

    css_string = f'''
        @font-face {{
            font-family: 'THSarabunNew';
            src: url('file://{font_path}') format('truetype');
            font-weight: normal;
        }}
        @font-face {{
            font-family: 'THSarabunNew';
            src: url('file://{font_bold_path}') format('truetype');
            font-weight: bold;
        }}
        @page {{
            size: A4;
            margin: 1.5cm 1cm;
        }}
        * {{
            font-family: 'THSarabunNew', 'DejaVu Sans', sans-serif;
        }}
        body {{
            font-size: 14pt;
            line-height: 1.4;
        }}
        .header {{
            text-align: center;
            margin-bottom: 0.5cm;
            border-bottom: 3px solid {badge_color};
            padding-bottom: 0.3cm;
        }}
        h1 {{
            font-size: 22pt;
            font-weight: bold;
            margin: 0.2cm 0;
            color: {badge_color};
        }}
        h2 {{
            font-size: 18pt;
            font-weight: bold;
            margin: 0.2cm 0;
        }}
        .meta {{
            font-size: 12pt;
            color: #666;
        }}
        .summary {{
            margin: 0.5cm 0;
            padding: 0.5cm;
            background-color: {badge_color}20;
            border: 2px solid {badge_color};
            border-radius: 8px;
        }}
        .summary-row {{
            display: flex;
            justify-content: space-around;
            margin: 0.3cm 0;
        }}
        .summary-item {{
            text-align: center;
        }}
        .summary-item h3 {{
            font-size: 28pt;
            font-weight: bold;
            margin: 0;
            color: {badge_color};
        }}
        .summary-item p {{
            font-size: 13pt;
            margin: 0.1cm 0;
            color: #333;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            margin-top: 0.5cm;
            font-size: 13pt;
        }}
        th, td {{
            border: 1px solid #000;
            padding: 0.2cm;
            text-align: left;
        }}
        th {{
            background-color: {badge_color};
            color: #000;
            font-weight: bold;
            text-align: center;
        }}
        .text-center {{
            text-align: center;
        }}
        .status-printed {{
            color: #10B981;
            font-weight: bold;
        }}
        .status-not-printed {{
            color: #F59E0B;
        }}
        .footer {{
            margin-top: 0.8cm;
            padding-top: 0.3cm;
            border-top: 1px solid #ccc;
            text-align: center;
            font-size: 11pt;
            color: #666;
        }}
    '''

    html = HTML(string=html_string, base_url=request.build_absolute_uri())
    pdf_file = html.write_pdf(
        stylesheets=[CSS(string=css_string, font_config=font_config)],
        font_config=font_config
    )

    # สร้างชื่อไฟล์ที่สื่อความหมาย
    filename = f"{badge_type.name}_{department.name}.pdf"

    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="{filename}"'

    return response


@login_required
def department_staff_export_excel(request, department_id):
    """
    Export Excel รายชื่อบุคลากรตามหน่วยงาน
    ตามฟอร์แมตที่ทหารต้องการ (10 คอลัมน์)
    """
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    # ตรวจสอบสิทธิ์
    if not request.user.can_manage_all():
        return render(request, '403.html', status=403)

    # ดึงข้อมูลหน่วยงาน
    try:
        department = Department.objects.get(id=department_id, is_active=True)
    except Department.DoesNotExist:
        return HttpResponse('ไม่พบหน่วยงาน', status=404)

    # ดึงข้อมูลบุคลากรทั้งหมดในหน่วยงาน
    staff_list = StaffProfile.objects.filter(
        department=department
    ).select_related(
        'department',
        'badge_type',
        'zone',
        'badge'
    ).order_by('last_line', 'first_line')

    # สร้าง Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "รายชื่อบุคลากร"

    # สไตล์สำหรับหัวตาราง
    header_font = Font(name='TH SarabunPSK', size=14, bold=True)
    header_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # สไตล์สำหรับข้อมูล
    data_font = Font(name='TH SarabunPSK', size=14)
    data_alignment_center = Alignment(horizontal='center', vertical='center')
    data_alignment_left = Alignment(horizontal='left', vertical='center')

    # เส้นขอบ
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # หัวตาราง (12 คอลัมน์)
    headers = [
        'ลำดับ',
        'ยศ - ชื่อ ยศ',
        'บัตรประชาชน 13 หลัก',
        'หน่วยงาน',
        'ประเภทบคคล',
        'ประเภทบัตร',
        'เลขที่บัตร',
        'บทบาทหน้าที่',
        'อายุ',
        'ฉาย',
        'พะเบียนรถ',
        'เบอร์โทรศัพท์มือถือ'
    ]

    # เขียนหัวตาราง
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # ตั้งค่าความกว้างคอลัมน์
    column_widths = [8, 30, 18, 35, 25, 20, 15, 25, 8, 15, 15, 18]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # เขียนข้อมูล
    row_num = 2
    for idx, staff in enumerate(staff_list, 1):
        # 1. ลำดับ
        cell = ws.cell(row=row_num, column=1)
        cell.value = idx
        cell.font = data_font
        cell.alignment = data_alignment_center
        cell.border = thin_border

        # 2. ยศ - ชื่อ ยศ
        cell = ws.cell(row=row_num, column=2)
        cell.value = staff.full_name
        cell.font = data_font
        cell.alignment = data_alignment_left
        cell.border = thin_border

        # 3. บัตรประชาชน 13 หลัก
        cell = ws.cell(row=row_num, column=3)
        cell.value = staff.national_id if staff.national_id else ''
        cell.font = data_font
        cell.alignment = data_alignment_center
        cell.border = thin_border
        # Format เป็นข้อความเพื่อไม่ให้เป็นเลข
        cell.number_format = '@'

        # 4. หน่วยงาน
        cell = ws.cell(row=row_num, column=4)
        cell.value = staff.department.name
        cell.font = data_font
        cell.alignment = data_alignment_left
        cell.border = thin_border

        # 5. ประเภทบคคล
        cell = ws.cell(row=row_num, column=5)
        cell.value = staff.person_type if staff.person_type else ''
        cell.font = data_font
        cell.alignment = data_alignment_left
        cell.border = thin_border

        # 6. ประเภทบัตร
        cell = ws.cell(row=row_num, column=6)
        cell.value = staff.badge_type.name if staff.badge_type else ''
        cell.font = data_font
        cell.alignment = data_alignment_center
        cell.border = thin_border

        # 7. เลขที่บัตร
        cell = ws.cell(row=row_num, column=7)
        if hasattr(staff, 'badge') and staff.badge:
            cell.value = staff.badge.badge_number
        else:
            cell.value = ''
        cell.font = data_font
        cell.alignment = data_alignment_center
        cell.border = thin_border

        # 8. บทบาทหน้าที่
        cell = ws.cell(row=row_num, column=8)
        cell.value = staff.position
        cell.font = data_font
        cell.alignment = data_alignment_left
        cell.border = thin_border

        # 9. อายุ
        cell = ws.cell(row=row_num, column=9)
        cell.value = staff.age if staff.age else ''
        cell.font = data_font
        cell.alignment = data_alignment_center
        cell.border = thin_border

        # 10. ฉาย (โซน)
        cell = ws.cell(row=row_num, column=10)
        cell.value = staff.zone.code if staff.zone else ''
        cell.font = data_font
        cell.alignment = data_alignment_center
        cell.border = thin_border

        # 11. พะเบียนรถ
        cell = ws.cell(row=row_num, column=11)
        cell.value = staff.vehicle_registration if staff.vehicle_registration else ''
        cell.font = data_font
        cell.alignment = data_alignment_center
        cell.border = thin_border

        # 12. เบอร์โทรศัพท์มือถือ
        cell = ws.cell(row=row_num, column=12)
        cell.value = staff.phone if staff.phone else ''
        cell.font = data_font
        cell.alignment = data_alignment_center
        cell.border = thin_border
        cell.number_format = '@'

        row_num += 1

    # Freeze หัวตาราง
    ws.freeze_panes = 'A2'

    # สร้าง response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'staff_list_{department.name}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # บันทึก workbook ลง response
    wb.save(response)

    return response
