#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Extract Images from Single Excel File
สคริปต์สำหรับ extract รูปจากไฟล์ Excel เดียว (แนะนำสำหรับผู้ใช้ทั่วไป)
"""

from openpyxl import load_workbook
import pandas as pd
import os
import re
from pathlib import Path

def extract_name_only(full_name):
    """ตัดยศออกจากชื่อ-สกุล"""
    if pd.isna(full_name):
        return None
    full_name = str(full_name).strip()

    # ยศทั่วไป + ยศทหาร/ตำรวจ (เรียงจากยาวไปสั้น)
    titles = [
        'นางสาว', 'เด็กชาย', 'เด็กหญิง',
        'พ.อ.', 'พ.ท.', 'พ.ต.', 'ร.อ.', 'ร.ท.', 'ร.ต.',
        'น.อ.', 'น.ท.', 'น.ต.', 'จ.อ.', 'จ.ท.', 'จ.ต.',
        'พล.อ.', 'พล.ท.', 'พล.ต.', 'พล.ร.อ.', 'พล.ร.ท.', 'พล.ร.ต.',
        'ด.ต.', 'ด.ท.', 'จ.ส.ต.', 'จ.ส.ท.', 'ส.ต.', 'ส.ท.',
        'พ.ต.อ.', 'พ.ต.ท.', 'พ.ต.ต.', 'ร.ต.อ.', 'ร.ต.ท.', 'ร.ต.ต.',
        'น.ส.ต.', 'น.ส.อ.', 'น.ส.ท.', 'ส.อ.', 'จ.ส.อ.',
        'น.สพ.', 'ด.ช.', 'ด.ญ.', 'คุณ', 'ท.', 'ต.', 'อ.',
        'นาย', 'นาง'
    ]

    for title in titles:
        if full_name.startswith(title):
            return full_name[len(title):].strip()

    return full_name


def extract_images_from_sheet(excel_file, sheet_name, output_folder, header_row=2):
    """Extract รูปภาพจาก 1 sheet"""

    os.makedirs(output_folder, exist_ok=True)

    # อ่านข้อมูล
    df = pd.read_excel(excel_file, sheet_name=sheet_name, header=header_row)

    # ตรวจสอบรูปแบบคอลัมน์
    has_separate_cols = 'ชื่อ' in df.columns and 'นามสกุล' in df.columns

    print(f'📊 ข้อมูล: {len(df)} แถว')
    print(f'📝 รูปแบบ: {"ชื่อ+นามสกุลแยก" if has_separate_cols else "รวมกัน"}')

    # Load workbook
    wb = load_workbook(excel_file)
    ws = wb[sheet_name]

    images_data = []
    for idx, image in enumerate(ws._images):
        row_num = image.anchor._from.row + 1
        images_data.append({'index': idx, 'row': row_num, 'image': image})

    print(f'🖼️  รูปภาพ: {len(images_data)} รูป\n')

    # Match และบันทึก
    saved_count = 0
    skipped_count = 0

    for img_data in images_data:
        df_index = img_data['row'] - header_row - 2

        if 0 <= df_index < len(df):
            record = df.iloc[df_index]

            # รวมชื่อ-นามสกุล
            if has_separate_cols:
                first_name = record['ชื่อ'] if 'ชื่อ' in record else None
                last_name = record['นามสกุล'] if 'นามสกุล' in record else None

                if pd.isna(first_name) or first_name == 'ชื่อ':
                    print(f'⊘ ข้าม row {img_data["row"]}: ไม่มีข้อมูลชื่อ')
                    skipped_count += 1
                    continue

                full_name = str(first_name).strip()
                if pd.notna(last_name):
                    full_name += ' ' + str(last_name).strip()
            else:
                # หาคอลัมน์ที่มีคำว่า 'ชื่อ'
                name_column = None
                for col in df.columns:
                    if 'ชื่อ' in str(col):
                        name_column = col
                        break

                if not name_column or pd.isna(record[name_column]):
                    print(f'⊘ ข้าม row {img_data["row"]}: ไม่มีข้อมูลชื่อ')
                    skipped_count += 1
                    continue

                full_name = str(record[name_column]).strip()

            if not full_name:
                print(f'⊘ ข้าม row {img_data["row"]}: ชื่อว่างเปล่า')
                skipped_count += 1
                continue

            name_only = extract_name_only(full_name)
            if not name_only:
                print(f'⊘ ข้าม row {img_data["row"]}: ไม่สามารถดึงชื่อได้')
                skipped_count += 1
                continue

            safe_name = re.sub(r'[<>:"/\\|?*]', '', name_only)
            base_filename = f'{safe_name}'
            filename = f'{base_filename}.png'

            # ถ้าไฟล์ซ้ำ ให้เพิ่มเลข
            counter = 1
            while os.path.exists(os.path.join(output_folder, filename)):
                filename = f'{base_filename}_{counter}.png'
                counter += 1

            filepath = os.path.join(output_folder, filename)

            with open(filepath, 'wb') as f:
                f.write(img_data['image']._data())

            print(f'✓ {filename}')
            saved_count += 1
        else:
            print(f'⚠ รูปที่ row {img_data["row"]} อยู่นอกช่วงข้อมูล')
            skipped_count += 1

    wb.close()

    return saved_count, skipped_count


def main():
    """Main function"""
    print('\n' + '='*70)
    print('📸 Extract Images from Excel - โปรแกรมดึงรูปจาก Excel')
    print('='*70 + '\n')

    # ตั้งค่า input/output folder
    input_folder = 'media/files/card_original'
    output_base_folder = 'media/photos/extracted'

    # หารายการไฟล์
    excel_files = list(Path(input_folder).glob('*.xlsx'))
    excel_files = [f for f in excel_files if not f.name.startswith('~$')]

    if not excel_files:
        print('❌ ไม่พบไฟล์ Excel ในโฟลเดอร์')
        return

    # แสดงรายการไฟล์
    print('📁 รายการไฟล์ Excel:')
    print('-' * 70)
    for idx, file in enumerate(excel_files, 1):
        print(f'  [{idx}] {file.name}')
    print()

    # เลือกไฟล์
    while True:
        try:
            choice = input('👉 เลือกไฟล์ (ระบุหมายเลข): ').strip()
            file_idx = int(choice) - 1

            if 0 <= file_idx < len(excel_files):
                selected_file = excel_files[file_idx]
                break
            else:
                print('❌ หมายเลขไม่ถูกต้อง กรุณาลองใหม่\n')
        except ValueError:
            print('❌ กรุณาระบุตัวเลข\n')

    print(f'\n✅ เลือกไฟล์: {selected_file.name}\n')

    # แสดงรายการ sheet
    try:
        wb = load_workbook(str(selected_file))
        sheets = wb.sheetnames

        print('📋 รายการ Sheet:')
        print('-' * 70)

        for idx, sheet in enumerate(sheets, 1):
            # แสดงจำนวนรูปใน sheet
            ws = wb[sheet]
            num_images = len(ws._images)
            print(f'  [{idx}] {sheet} ({num_images} รูป)')

        wb.close()
        print()

    except Exception as e:
        print(f'❌ Error ในการอ่านไฟล์: {e}')
        return

    # เลือก sheet
    while True:
        choice = input('👉 เลือก Sheet (ระบุหมายเลข หรือ 0 สำหรับทุก sheet): ').strip()

        try:
            if choice == '0':
                selected_sheets = sheets
                print(f'\n✅ จะประมวลผลทุก sheet ({len(sheets)} sheets)\n')
                break
            else:
                sheet_idx = int(choice) - 1
                if 0 <= sheet_idx < len(sheets):
                    selected_sheets = [sheets[sheet_idx]]
                    print(f'\n✅ เลือก sheet: {sheets[sheet_idx]}\n')
                    break
                else:
                    print('❌ หมายเลขไม่ถูกต้อง\n')
        except ValueError:
            print('❌ กรุณาระบุตัวเลข\n')

    # ประมวลผล
    print('='*70)
    print('🚀 เริ่มประมวลผล...')
    print('='*70 + '\n')

    total_saved = 0
    total_skipped = 0

    for sheet in selected_sheets:
        print(f'📄 Sheet: {sheet}')
        print('-' * 70)

        # สร้าง output folder
        file_name_clean = selected_file.stem
        output_folder = os.path.join(output_base_folder, file_name_clean, sheet)

        try:
            saved, skipped = extract_images_from_sheet(
                excel_file=str(selected_file),
                sheet_name=sheet,
                output_folder=output_folder,
                header_row=2
            )
            total_saved += saved
            total_skipped += skipped

            print(f'📌 สรุป: บันทึก {saved} รูป | ข้าม {skipped} รูป\n')

        except Exception as e:
            print(f'❌ Error: {e}\n')

    print('='*70)
    print(f'✅ เสร็จสิ้น!')
    print(f'📊 บันทึกทั้งหมด: {total_saved} รูป | ข้าม: {total_skipped} รูป')
    print(f'📁 ที่ตำแหน่ง: {output_base_folder}/{file_name_clean}/')
    print('='*70 + '\n')


if __name__ == '__main__':
    main()
