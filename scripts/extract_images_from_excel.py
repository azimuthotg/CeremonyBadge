from openpyxl import load_workbook
import pandas as pd
import os
from pathlib import Path
import re

def extract_name_only(full_name):
    """
    ตัดยศออกจากชื่อ-สกุล
    เช่น 'นายอลงกรณ์ ดอกจันรี' -> 'อลงกรณ์ ดอกจันรี'
    """
    if pd.isna(full_name):
        return None

    full_name = str(full_name).strip()

    # ยศที่ต้องการตัดออก (เรียงจากยาวไปสั้น เพื่อให้จับคำที่ยาวกว่าก่อน)
    titles = [
        'นางสาว', 'เด็กชาย', 'เด็กหญิง',
        'พ.อ.', 'พ.ท.', 'พ.ต.', 'ร.อ.', 'ร.ท.', 'ร.ต.',
        'น.อ.', 'น.ท.', 'น.ต.', 'จ.อ.', 'จ.ท.', 'จ.ต.',
        'พล.อ.', 'พล.ท.', 'พล.ต.', 'พล.ร.อ.', 'พล.ร.ท.', 'พล.ร.ต.',
        'ด.ต.', 'ด.ท.', 'จ.ส.ต.', 'จ.ส.ท.', 'ส.ต.', 'ส.ท.',
        'พ.ต.อ.', 'พ.ต.ท.', 'พ.ต.ต.', 'ร.ต.อ.', 'ร.ต.ท.', 'ร.ต.ต.',
        'น.ส.ต.', 'น.ส.อ.', 'น.ส.ท.', 'ส.อ.', 'จ.ส.อ.',
        'น.สพ.', 'ด.ช.', 'ด.ญ.', 'คุณ', 'ท.', 'ต.', 'อ.',
        'นาย', 'นาง'
    ]

    for title in titles:
        if full_name.startswith(title):
            return full_name[len(title):].strip()

    return full_name


def extract_images_with_names(excel_file, output_folder='exported_images',
                               name_column='ยศ - ชื่อ - สกุล',
                               sheet_name=None, header_row=2):
    """
    Extract รูปภาพจาก Excel และตั้งชื่อตามชื่อ-สกุล

    Args:
        excel_file: path ไฟล์ Excel
        output_folder: โฟลเดอร์สำหรับเก็บรูป
        name_column: คอลัมน์ที่ใช้ตั้งชื่อไฟล์
        sheet_name: sheet ที่จะ extract (None = ทุก sheet)
        header_row: แถวที่เป็น header (เริ่มนับจาก 0)
    """

    # สร้างโฟลเดอร์
    os.makedirs(output_folder, exist_ok=True)

    # Load workbook
    wb = load_workbook(excel_file)

    # ถ้าไม่ระบุ sheet ให้ทำทุก sheet
    sheets_to_process = [sheet_name] if sheet_name else wb.sheetnames

    total_saved = 0
    total_errors = []

    for current_sheet in sheets_to_process:
        print(f"\n{'='*60}")
        print(f"กำลังประมวลผล Sheet: {current_sheet}")
        print(f"{'='*60}")

        # 1. อ่านข้อมูลจาก Excel
        df = pd.read_excel(excel_file, sheet_name=current_sheet, header=header_row)
        print(f"พบข้อมูล {len(df)} แถว")

        # 2. ดึงรูปภาพ
        ws = wb[current_sheet]
        images_data = []

        for idx, image in enumerate(ws._images):
            try:
                row_num = image.anchor._from.row + 1
                col_num = image.anchor._from.col

                images_data.append({
                    'index': idx,
                    'row': row_num,
                    'col': col_num,
                    'image': image
                })
            except Exception as e:
                print(f"ข้าม image {idx}: {e}")

        print(f"พบรูปภาพ {len(images_data)} รูป")

        # 3. Match รูปกับข้อมูล และบันทึก
        saved_count = 0
        errors = []

        for img_data in images_data:
            try:
                # คำนวณ index ของ DataFrame
                df_index = img_data['row'] - header_row - 2

                if 0 <= df_index < len(df):
                    record = df.iloc[df_index]

                    # เช็คว่ามีข้อมูลชื่อหรือไม่
                    if name_column not in record or pd.isna(record[name_column]):
                        print(f"⊘ ข้าม row {img_data['row']}: ไม่มีข้อมูลชื่อ")
                        continue

                    # ดึงชื่อ-สกุล และตัดยศออก
                    full_name = str(record[name_column]).strip()
                    if not full_name:
                        print(f"⊘ ข้าม row {img_data['row']}: ชื่อว่างเปล่า")
                        continue

                    name_only = extract_name_only(full_name)
                    if not name_only:
                        print(f"⊘ ข้าม row {img_data['row']}: ไม่สามารถดึงชื่อได้")
                        continue

                    # ลบอักขระพิเศษที่ใช้ในชื่อไฟล์ไม่ได้
                    safe_name = re.sub(r'[<>:"/\\|?*]', '', name_only)
                    filename = f"{safe_name}.png"

                    # สร้าง subfolder สำหรับแต่ละ sheet
                    sheet_folder = os.path.join(output_folder, current_sheet)
                    os.makedirs(sheet_folder, exist_ok=True)

                    # บันทึกรูป
                    filepath = os.path.join(sheet_folder, filename)

                    with open(filepath, 'wb') as f:
                        f.write(img_data['image']._data())

                    print(f"✓ บันทึก: {filename}")
                    saved_count += 1

                else:
                    error_msg = f"รูปที่ row {img_data['row']} อยู่นอกช่วงข้อมูล (df_index={df_index})"
                    print(f"⚠ {error_msg}")
                    errors.append(error_msg)

            except Exception as e:
                error_msg = f"Error ที่ row {img_data['row']}: {str(e)}"
                print(f"✗ {error_msg}")
                errors.append(error_msg)

        print(f"\nSheet '{current_sheet}': บันทึกรูปสำเร็จ {saved_count}/{len(images_data)} รูป")
        total_saved += saved_count
        total_errors.extend(errors)

    wb.close()
    return total_saved, total_errors


def process_all_excel_files(input_folder, output_folder):
    """
    ประมวลผลไฟล์ Excel ทั้งหมดใน folder
    """
    print(f"\n{'#'*70}")
    print(f"# เริ่มประมวลผลไฟล์ Excel ทั้งหมดใน {input_folder}")
    print(f"{'#'*70}\n")

    excel_files = list(Path(input_folder).glob("*.xlsx"))
    excel_files = [f for f in excel_files if not f.name.startswith('~$')]  # ข้ามไฟล์ temp

    print(f"พบไฟล์ Excel {len(excel_files)} ไฟล์\n")

    total_images = 0

    for idx, excel_file in enumerate(excel_files, 1):
        print(f"\n{'='*70}")
        print(f"[{idx}/{len(excel_files)}] ไฟล์: {excel_file.name}")
        print(f"{'='*70}")

        try:
            saved, errors = extract_images_with_names(
                excel_file=str(excel_file),
                output_folder=output_folder,
                name_column='ยศ - ชื่อ - สกุล',
                sheet_name=None,  # ทำทุก sheet
                header_row=2
            )
            total_images += saved

            if errors:
                print(f"\n⚠ พบปัญหา {len(errors)} รายการในไฟล์นี้")

        except Exception as e:
            print(f"\n✗ Error ในการประมวลผลไฟล์: {e}")

    print(f"\n{'#'*70}")
    print(f"# สรุปผลรวม: บันทึกรูปสำเร็จทั้งหมด {total_images} รูป")
    print(f"{'#'*70}\n")


# ===== ตัวอย่างการใช้งาน =====
if __name__ == "__main__":
    # วิธีที่ 1: ประมวลผลไฟล์เดียว
    # saved, errors = extract_images_with_names(
    #     excel_file="media/files/card_original/4. ท่าอากาศยานนครพนม.xlsx",
    #     output_folder='media/photos/extracted',
    #     name_column='ยศ - ชื่อ - สกุล',
    #     sheet_name=None,  # ทุก sheet
    #     header_row=2
    # )

    # วิธีที่ 2: ประมวลผลทุกไฟล์ใน folder
    process_all_excel_files(
        input_folder='media/files/card_original',
        output_folder='media/photos/extracted'
    )
