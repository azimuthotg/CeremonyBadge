from openpyxl import load_workbook
import pandas as pd
import os
from pathlib import Path
import re

def extract_name_only(full_name):
    """
    ตัดยศออกจากชื่อ-สกุล
    เช่น 'นายอลงกรณ์ ดอกจันรี' -> 'อลงกรณ์ ดอกจันรี'
    """
    if pd.isna(full_name):
        return None

    full_name = str(full_name).strip()

    # ยศที่ต้องการตัดออก (เรียงจากยาวไปสั้น เพื่อให้จับคำที่ยาวกว่าก่อน)
    titles = [
        'นางสาว', 'เด็กชาย', 'เด็กหญิง',
        'พ.อ.', 'พ.ท.', 'พ.ต.', 'ร.อ.', 'ร.ท.', 'ร.ต.',
        'น.อ.', 'น.ท.', 'น.ต.', 'จ.อ.', 'จ.ท.', 'จ.ต.',
        'พล.อ.', 'พล.ท.', 'พล.ต.', 'พล.ร.อ.', 'พล.ร.ท.', 'พล.ร.ต.',
        'ด.ต.', 'ด.ท.', 'จ.ส.ต.', 'จ.ส.ท.', 'ส.ต.', 'ส.ท.',
        'พ.ต.อ.', 'พ.ต.ท.', 'พ.ต.ต.', 'ร.ต.อ.', 'ร.ต.ท.', 'ร.ต.ต.',
        'น.ส.ต.', 'น.ส.อ.', 'น.ส.ท.', 'ส.อ.', 'จ.ส.อ.',
        'น.สพ.', 'ด.ช.', 'ด.ญ.', 'คุณ', 'ท.', 'ต.', 'อ.',
        'นาย', 'นาง'
    ]

    for title in titles:
        if full_name.startswith(title):
            return full_name[len(title):].strip()

    return full_name


def extract_images_from_sheet(excel_file, sheet_name, output_folder, name_column='ยศ - ชื่อ - สกุล', header_row=2):
    """
    Extract รูปภาพจาก 1 sheet
    """
    # สร้างโฟลเดอร์
    os.makedirs(output_folder, exist_ok=True)

    # 1. อ่านข้อมูลจาก Excel
    df = pd.read_excel(excel_file, sheet_name=sheet_name, header=header_row)
    print(f"  📊 พบข้อมูล {len(df)} แถว")

    # 2. Load workbook และดึงรูปภาพ
    wb = load_workbook(excel_file)
    ws = wb[sheet_name]

    images_data = []
    for idx, image in enumerate(ws._images):
        try:
            row_num = image.anchor._from.row + 1
            col_num = image.anchor._from.col

            images_data.append({
                'index': idx,
                'row': row_num,
                'col': col_num,
                'image': image
            })
        except Exception as e:
            print(f"  ⚠ ข้าม image {idx}: {e}")

    print(f"  🖼️  พบรูปภาพ {len(images_data)} รูป")

    # 3. Match รูปกับข้อมูล และบันทึก
    saved_count = 0
    errors = []

    for img_data in images_data:
        try:
            # คำนวณ index ของ DataFrame
            df_index = img_data['row'] - header_row - 2

            if 0 <= df_index < len(df):
                record = df.iloc[df_index]

                # เช็คว่ามีข้อมูลชื่อหรือไม่
                if name_column not in record or pd.isna(record[name_column]):
                    print(f"  ⊘ ข้าม row {img_data['row']}: ไม่มีข้อมูลชื่อ")
                    continue

                # ดึงชื่อ-สกุล และตัดยศออก
                full_name = str(record[name_column]).strip()
                if not full_name:
                    print(f"  ⊘ ข้าม row {img_data['row']}: ชื่อว่างเปล่า")
                    continue

                name_only = extract_name_only(full_name)
                if not name_only:
                    print(f"  ⊘ ข้าม row {img_data['row']}: ไม่สามารถดึงชื่อได้")
                    continue

                # ลบอักขระพิเศษ
                safe_name = re.sub(r'[<>:"/\\|?*]', '', name_only)
                filename = f"{safe_name}.png"

                # บันทึกรูป
                filepath = os.path.join(output_folder, filename)

                with open(filepath, 'wb') as f:
                    f.write(img_data['image']._data())

                print(f"  ✓ บันทึก: {filename}")
                saved_count += 1

            else:
                error_msg = f"รูปที่ row {img_data['row']} อยู่นอกช่วงข้อมูล"
                print(f"  ⚠ {error_msg}")
                errors.append(error_msg)

        except Exception as e:
            error_msg = f"Error ที่ row {img_data['row']}: {str(e)}"
            print(f"  ✗ {error_msg}")
            errors.append(error_msg)

    wb.close()

    print(f"\n  📌 สรุป: บันทึกสำเร็จ {saved_count}/{len(images_data)} รูป")
    if errors:
        print(f"  ⚠ พบปัญหา {len(errors)} รายการ")

    return saved_count, errors


def interactive_mode():
    """
    โหมดเลือกไฟล์และ sheet แบบ interactive
    """
    print("\n" + "="*70)
    print(" 🎯 โปรแกรม Extract รูปภาพจาก Excel")
    print("="*70 + "\n")

    # ตั้งค่า input/output folder
    input_folder = 'media/files/card_original'
    output_base_folder = 'media/photos/extracted'

    while True:
        # 1. แสดงรายการไฟล์
        print("\n" + "-"*70)
        print("📁 รายการไฟล์ Excel:")
        print("-"*70)

        excel_files = list(Path(input_folder).glob("*.xlsx"))
        excel_files = [f for f in excel_files if not f.name.startswith('~$')]

        if not excel_files:
            print("❌ ไม่พบไฟล์ Excel ในโฟลเดอร์")
            break

        for idx, file in enumerate(excel_files, 1):
            print(f"  [{idx}] {file.name}")

        print(f"  [0] ออกจากโปรแกรม")

        # 2. เลือกไฟล์
        try:
            file_choice = input("\n👉 เลือกไฟล์ (ระบุหหมายเลข): ").strip()

            if file_choice == '0':
                print("\n✅ ออกจากโปรแกรม")
                break

            file_idx = int(file_choice) - 1
            if file_idx < 0 or file_idx >= len(excel_files):
                print("❌ หมายเลขไม่ถูกต้อง กรุณาลองใหม่")
                continue

            selected_file = excel_files[file_idx]

        except ValueError:
            print("❌ กรุณาระบุตัวเลข")
            continue

        print(f"\n✅ เลือกไฟล์: {selected_file.name}")

        # 3. แสดงรายการ sheet
        try:
            wb = load_workbook(str(selected_file))
            sheets = wb.sheetnames
            wb.close()

            print("\n" + "-"*70)
            print("📋 รายการ Sheet:")
            print("-"*70)

            for idx, sheet in enumerate(sheets, 1):
                # แสดงจำนวนรูปใน sheet
                wb_temp = load_workbook(str(selected_file))
                ws_temp = wb_temp[sheet]
                num_images = len(ws_temp._images)
                wb_temp.close()
                print(f"  [{idx}] {sheet} ({num_images} รูป)")

            print(f"  [0] ทำทุก sheet")
            print(f"  [x] ยกเลิก กลับไปเลือกไฟล์ใหม่")

        except Exception as e:
            print(f"❌ Error ในการอ่านไฟล์: {e}")
            continue

        # 4. เลือก sheet
        sheet_choice = input("\n👉 เลือก Sheet (ระบุหหมายเลข): ").strip()

        if sheet_choice.lower() == 'x':
            continue

        if sheet_choice == '0':
            # ทำทุก sheet
            selected_sheets = sheets
            print(f"\n✅ จะประมวลผลทุก sheet ({len(sheets)} sheets)")
        else:
            try:
                sheet_idx = int(sheet_choice) - 1
                if sheet_idx < 0 or sheet_idx >= len(sheets):
                    print("❌ หมายเลขไม่ถูกต้อง")
                    continue
                selected_sheets = [sheets[sheet_idx]]
                print(f"\n✅ เลือก sheet: {sheets[sheet_idx]}")
            except ValueError:
                print("❌ กรุณาระบุตัวเลข")
                continue

        # 5. ประมวลผล
        print("\n" + "="*70)
        print("🚀 เริ่มประมวลผล...")
        print("="*70)

        total_saved = 0

        for sheet in selected_sheets:
            print(f"\n📄 Sheet: {sheet}")
            print("-"*70)

            # สร้าง output folder แยกตามไฟล์และ sheet
            file_name_clean = selected_file.stem  # ชื่อไฟล์ไม่มี .xlsx
            output_folder = os.path.join(output_base_folder, file_name_clean, sheet)

            try:
                saved, errors = extract_images_from_sheet(
                    excel_file=str(selected_file),
                    sheet_name=sheet,
                    output_folder=output_folder,
                    name_column='ยศ - ชื่อ - สกุล',
                    header_row=2
                )
                total_saved += saved

            except Exception as e:
                print(f"  ❌ Error: {e}")

        print("\n" + "="*70)
        print(f"✅ เสร็จสิ้น! บันทึกรูปทั้งหมด {total_saved} รูป")
        print(f"📁 ที่ตำแหน่ง: {output_base_folder}/{file_name_clean}/")
        print("="*70)

        # 6. ถามว่าจะทำต่อไหม
        continue_choice = input("\n👉 ต้องการประมวลผลไฟล์อื่นต่อไหม? (y/n): ").strip().lower()
        if continue_choice != 'y':
            print("\n✅ เสร็จสิ้นทั้งหมด ขอบคุณครับ!")
            break


if __name__ == "__main__":
    interactive_mode()
