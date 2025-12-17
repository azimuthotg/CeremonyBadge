#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Auto-extract images from file 16.บัตรบุคคลมหาวิทยาลัยนครพนม.xlsx
"""

from openpyxl import load_workbook
import pandas as pd
import os
import re
from pathlib import Path

def extract_name_only(full_name):
    """ตัดยศออกจากชื่อ-สกุล"""
    if pd.isna(full_name):
        return None
    full_name = str(full_name).strip()

    # ยศทั่วไป + ยศทหาร/ตำรวจ (เรียงจากยาวไปสั้น)
    titles = [
        'นางสาว', 'เด็กชาย', 'เด็กหญิง',
        'พ.อ.', 'พ.ท.', 'พ.ต.', 'ร.อ.', 'ร.ท.', 'ร.ต.',
        'น.อ.', 'น.ท.', 'น.ต.', 'จ.อ.', 'จ.ท.', 'จ.ต.',
        'พล.อ.', 'พล.ท.', 'พล.ต.', 'พล.ร.อ.', 'พล.ร.ท.', 'พล.ร.ต.',
        'ด.ต.', 'ด.ท.', 'จ.ส.ต.', 'จ.ส.ท.', 'ส.ต.', 'ส.ท.',
        'พ.ต.อ.', 'พ.ต.ท.', 'พ.ต.ต.', 'ร.ต.อ.', 'ร.ต.ท.', 'ร.ต.ต.',
        'น.ส.ต.', 'น.ส.อ.', 'น.ส.ท.', 'ส.อ.', 'จ.ส.อ.',
        'น.สพ.', 'ด.ช.', 'ด.ญ.', 'คุณ', 'ท.', 'ต.', 'อ.',
        'ศ.ดร.', 'รศ.ดร.', 'ผศ.ดร.', 'ดร.', 'ศ.', 'รศ.', 'ผศ.',
        'นาย', 'นาง'
    ]

    for title in titles:
        if full_name.startswith(title):
            return full_name[len(title):].strip()

    return full_name


def extract_images_from_sheet(excel_file, sheet_name, output_folder, header_row=2):
    """Extract รูปภาพจาก 1 sheet"""

    os.makedirs(output_folder, exist_ok=True)

    # อ่านข้อมูล
    df = pd.read_excel(excel_file, sheet_name=sheet_name, header=header_row)

    # ตรวจสอบรูปแบบคอลัมน์
    has_separate_cols = 'ชื่อ' in df.columns and 'นามสกุล' in df.columns

    # โหลด workbook เพื่อดึงรูป
    wb = load_workbook(excel_file)
    ws = wb[sheet_name]

    # ดึงรูปภาพทั้งหมด + ตำแหน่งของมัน
    images_data = []
    for idx, image in enumerate(ws._images):
        row_num = image.anchor._from.row + 1  # Convert to 1-based
        col_num = image.anchor._from.col + 1
        images_data.append({
            'index': idx,
            'row': row_num,
            'col': col_num,
            'image': image
        })

    print(f"\n📊 Sheet: {sheet_name}")
    print(f"   - พบรูปภาพ {len(images_data)} รูป")
    print(f"   - พบข้อมูล {len(df)} แถว")

    # บันทึกรูปภาพ
    saved_count = 0
    skipped_count = 0

    for img_data in images_data:
        # คำนวณ index ของแถวใน DataFrame
        df_index = img_data['row'] - header_row - 2

        if 0 <= df_index < len(df):
            record = df.iloc[df_index]

            # สร้างชื่อไฟล์
            if has_separate_cols:
                rank = record.get('ยศ', '')
                first = record.get('ชื่อ', '')
                last = record.get('นามสกุล', '')

                if pd.notna(rank) and str(rank).strip():
                    full_name = f"{rank}{first} {last}"
                else:
                    full_name = f"{first} {last}"
            else:
                # ใช้คอลัมน์ที่มีคำว่า "ชื่อ" หรือ "นามสกุล"
                name_col = None
                for col in df.columns:
                    if 'ชื่อ' in str(col) or 'นามสกุล' in str(col):
                        name_col = col
                        break

                if name_col:
                    full_name = str(record[name_col])
                else:
                    full_name = f"image_{img_data['index'] + 1}"

            # ตัดยศออก
            name_only = extract_name_only(full_name)

            if name_only:
                # ลบอักขระพิเศษ
                safe_name = re.sub(r'[<>:"/\\|?*]', '', name_only)
                safe_name = safe_name.strip()

                # ตรวจสอบว่าซ้ำไหม
                base_filename = safe_name
                counter = 1
                filename = f'{base_filename}.png'
                filepath = os.path.join(output_folder, filename)

                while os.path.exists(filepath):
                    filename = f'{base_filename}_{counter}.png'
                    filepath = os.path.join(output_folder, filename)
                    counter += 1

                # บันทึกรูป
                try:
                    with open(filepath, 'wb') as f:
                        f.write(img_data['image']._data())
                    print(f"   ✅ {filename}")
                    saved_count += 1
                except Exception as e:
                    print(f"   ❌ Error saving {filename}: {e}")
                    skipped_count += 1
            else:
                print(f"   ⚠️  ข้ามรูปที่ row {img_data['row']} (ไม่มีชื่อ)")
                skipped_count += 1
        else:
            print(f"   ⚠️  ข้ามรูปที่ row {img_data['row']} (นอกช่วงข้อมูล)")
            skipped_count += 1

    print(f"   ✅ บันทึกสำเร็จ: {saved_count} รูป")
    if skipped_count > 0:
        print(f"   ⚠️  ข้าม: {skipped_count} รูป")

    wb.close()
    return saved_count


def main():
    print("="*70)
    print("📸 Auto Extract Images from Excel File #16")
    print("="*70)

    # กำหนดพาธ
    input_folder = 'media/files/card_original'
    output_base_folder = 'media/photos/extracted'
    filename = '16.บัตรบุคคลมหาวิทยาลัยนครพนม.xlsx'

    excel_file = os.path.join(input_folder, filename)

    # ตรวจสอบว่าไฟล์มีอยู่
    if not os.path.exists(excel_file):
        print(f"❌ ไม่พบไฟล์: {excel_file}")
        return

    print(f"\n📁 กำลังประมวลผลไฟล์: {filename}")

    # โหลด workbook เพื่อดึงชื่อ sheet
    wb = load_workbook(excel_file, read_only=True)
    sheet_names = wb.sheetnames
    wb.close()

    print(f"\n📋 พบ {len(sheet_names)} sheet:")
    for i, sheet in enumerate(sheet_names, 1):
        print(f"   [{i}] {sheet}")

    # สร้างโฟลเดอร์สำหรับไฟล์นี้
    file_name_clean = re.sub(r'[<>:"/\\|?*]', '', filename.replace('.xlsx', ''))

    total_saved = 0

    # ประมวลผลทุก sheet
    print(f"\n🚀 เริ่มต้น extract รูปภาพ...\n")

    for sheet in sheet_names:
        # สร้างโฟลเดอร์ output
        output_folder = os.path.join(output_base_folder, file_name_clean, sheet)

        try:
            saved = extract_images_from_sheet(excel_file, sheet, output_folder, header_row=2)
            total_saved += saved
        except Exception as e:
            print(f"\n❌ Error processing sheet '{sheet}': {e}")
            import traceback
            traceback.print_exc()

    print("\n" + "="*70)
    print(f"✅ เสร็จสิ้น! บันทึกรูปภาพทั้งหมด {total_saved} รูป")
    print(f"📁 โฟลเดอร์: {os.path.join(output_base_folder, file_name_clean)}")
    print("="*70)


if __name__ == '__main__':
    main()
