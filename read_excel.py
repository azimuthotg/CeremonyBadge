import openpyxl

wb = openpyxl.load_workbook('media/files/test_import_pink.xlsx')
ws = wb.active

print('Sheet name:', ws.title)
print('\nHeaders:')
headers = [cell.value for cell in ws[1]]
print(headers)

print('\nFirst 10 rows:')
for i, row in enumerate(ws.iter_rows(min_row=2, max_row=11, values_only=True), start=2):
    print(f'Row {i}:', row)

print('\nTotal rows:', ws.max_row, 'Total columns:', ws.max_column)
